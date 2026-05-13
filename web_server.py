# web_server.py
from flask import Flask, render_template_string, request, jsonify, session, redirect, url_for
import secrets
import threading
import webbrowser
import pymysql
from datetime import datetime
from app_config import REGIONS_FULL, MYSQL_USER, MYSQL_PASSWORD, MYSQL_PORT
from user_manager import UserManager

web_app = Flask(__name__)
web_app.secret_key = secrets.token_hex(16)

# HTML шаблон с полным функционалом
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Поиск транзакций - Веб версия</title>
    <meta charset="utf-8">
    <style>
        * { box-sizing: border-box; }
        body { font-family: 'Segoe UI', Arial, sans-serif; margin: 0; padding: 20px; background: #f0f2f5; }
        .container { max-width: 1400px; margin: 0 auto; background: white; padding: 25px; border-radius: 12px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        .header { background: linear-gradient(135deg, #4a86e8, #3a76d8); color: white; padding: 20px; margin: -25px -25px 25px -25px; border-radius: 12px 12px 0 0; }
        .header h1 { margin: 0; font-size: 24px; }
        .header p { margin: 5px 0 0; opacity: 0.9; }
        .login-form { max-width: 400px; margin: 100px auto; background: white; padding: 30px; border-radius: 12px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        .login-form h2 { margin-top: 0; color: #4a86e8; }
        .form-group { margin-bottom: 15px; }
        label { display: block; margin-bottom: 5px; font-weight: bold; color: #333; }
        input, select { width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 6px; font-size: 14px; }
        input:focus, select:focus { outline: none; border-color: #4a86e8; }
        button { background: #4a86e8; color: white; border: none; padding: 10px 20px; border-radius: 6px; cursor: pointer; font-size: 14px; font-weight: bold; }
        button:hover { background: #3a76d8; }
        button.danger { background: #dc3545; }
        button.danger:hover { background: #c82333; }
        button.success { background: #28a745; }
        button.success:hover { background: #218838; }
        .search-section { background: #f8f9fa; padding: 20px; border-radius: 8px; margin-bottom: 20px; }
        .row { display: flex; gap: 20px; margin-bottom: 15px; flex-wrap: wrap; }
        .col { flex: 1; min-width: 200px; }
        .date-row { display: flex; gap: 10px; align-items: center; }
        .date-row select { width: auto; }
        .checkbox-group { display: flex; gap: 20px; margin-bottom: 15px; }
        .checkbox-group label { display: flex; align-items: center; gap: 5px; font-weight: normal; cursor: pointer; }
        .results-table { overflow-x: auto; margin-top: 20px; }
        table { width: 100%; border-collapse: collapse; font-size: 13px; }
        th, td { border: 1px solid #ddd; padding: 10px; text-align: left; }
        th { background: #4a86e8; color: white; position: sticky; top: 0; }
        tr:nth-child(even) { background: #f9f9f9; }
        tr:hover { background: #f0f0f0; cursor: pointer; }
        .loading { text-align: center; padding: 40px; color: #666; }
        .error { color: red; padding: 10px; background: #ffe0e0; border-radius: 6px; margin: 10px 0; }
        .info { color: #4a86e8; padding: 10px; background: #e0e8ff; border-radius: 6px; margin: 10px 0; }
        .btn-group { display: flex; gap: 10px; margin-top: 15px; flex-wrap: wrap; }
        .logout-btn { float: right; background: rgba(255,255,255,0.2); padding: 8px 15px; border-radius: 6px; text-decoration: none; color: white; }
        .logout-btn:hover { background: rgba(255,255,255,0.3); }
        .query-list { background: #f8f9fa; padding: 15px; border-radius: 8px; margin-bottom: 20px; }
        .query-item { display: inline-block; background: #e9ecef; padding: 8px 15px; margin: 5px; border-radius: 20px; cursor: pointer; font-size: 13px; }
        .query-item:hover { background: #4a86e8; color: white; }
        @media (max-width: 768px) { .row { flex-direction: column; } }
    </style>
</head>
<body>
    {% if not session.user %}
    <div class="login-form">
        <h2>Вход в систему</h2>
        {% if error %}<div class="error">{{ error }}</div>{% endif %}
        <form method="post">
            <div class="form-group">
                <label>Логин</label>
                <input type="text" name="username" required>
            </div>
            <div class="form-group">
                <label>Пароль</label>
                <input type="password" name="password" required>
            </div>
            <button type="submit">Войти</button>
        </form>
        <div style="margin-top: 15px; font-size: 12px; color: #666;">
            <hr>
            <p>Тестовый доступ: admin / admin123</p>
        </div>
    </div>
    {% else %}
    <div class="container">
        <div class="header">
            <h1>🔍 Поиск транзакций</h1>
            <p>Пользователь: {{ session.user.full_name }} ({{ session.user.username }})</p>
            <a href="/logout" class="logout-btn">Выйти</a>
        </div>
        
        <div class="search-section">
            <div class="checkbox-group">
                <label>
                    <input type="checkbox" id="searchByTransaction" checked onchange="toggleDateFrames()">
                    Поиск по дате регистрации
                </label>
                <label>
                    <input type="checkbox" id="searchByBankdate" onchange="toggleDateFrames()">
                    Поиск по дате списания
                </label>
            </div>
            
            <div id="transactionFrame" style="margin-bottom: 15px;">
                <h3>Дата и время регистрации на терминале</h3>
                <div class="row">
                    <div class="col">
                        <label>С</label>
                        <div class="date-row">
                            <input type="date" id="transDateFrom" value="{{ today }}">
                            <select id="transHourFrom">
                                {% for i in range(24) %}<option value="{{ "%02d"|format(i) }}" {% if i==0 %}selected{% endif %}>{{ "%02d"|format(i) }}</option>{% endfor %}
                            </select> :
                            <select id="transMinFrom">
                                {% for i in range(60) %}<option value="{{ "%02d"|format(i) }}" {% if i==0 %}selected{% endif %}>{{ "%02d"|format(i) }}</option>{% endfor %}
                            </select> :
                            <select id="transSecFrom">
                                {% for i in range(60) %}<option value="{{ "%02d"|format(i) }}" {% if i==0 %}selected{% endif %}>{{ "%02d"|format(i) }}</option>{% endfor %}
                            </select>
                        </div>
                    </div>
                    <div class="col">
                        <label>По</label>
                        <div class="date-row">
                            <input type="date" id="transDateTo" value="{{ today }}">
                            <select id="transHourTo">
                                {% for i in range(24) %}<option value="{{ "%02d"|format(i) }}" {% if i==23 %}selected{% endif %}>{{ "%02d"|format(i) }}</option>{% endfor %}
                            </select> :
                            <select id="transMinTo">
                                {% for i in range(60) %}<option value="{{ "%02d"|format(i) }}" {% if i==59 %}selected{% endif %}>{{ "%02d"|format(i) }}</option>{% endfor %}
                            </select> :
                            <select id="transSecTo">
                                {% for i in range(60) %}<option value="{{ "%02d"|format(i) }}" {% if i==59 %}selected{% endif %}>{{ "%02d"|format(i) }}</option>{% endfor %}
                            </select>
                        </div>
                    </div>
                </div>
            </div>
            
            <div id="bankdateFrame" style="display: none; margin-bottom: 15px;">
                <h3>Дата и время списания</h3>
                <div class="row">
                    <div class="col">
                        <label>С</label>
                        <div class="date-row">
                            <input type="date" id="bankDateFrom" value="{{ today }}">
                            <select id="bankHourFrom">
                                {% for i in range(24) %}<option value="{{ "%02d"|format(i) }}" {% if i==0 %}selected{% endif %}>{{ "%02d"|format(i) }}</option>{% endfor %}
                            </select> :
                            <select id="bankMinFrom">
                                {% for i in range(60) %}<option value="{{ "%02d"|format(i) }}" {% if i==0 %}selected{% endif %}>{{ "%02d"|format(i) }}</option>{% endfor %}
                            </select> :
                            <select id="bankSecFrom">
                                {% for i in range(60) %}<option value="{{ "%02d"|format(i) }}" {% if i==0 %}selected{% endif %}>{{ "%02d"|format(i) }}</option>{% endfor %}
                            </select>
                        </div>
                    </div>
                    <div class="col">
                        <label>По</label>
                        <div class="date-row">
                            <input type="date" id="bankDateTo" value="{{ today }}">
                            <select id="bankHourTo">
                                {% for i in range(24) %}<option value="{{ "%02d"|format(i) }}" {% if i==23 %}selected{% endif %}>{{ "%02d"|format(i) }}</option>{% endfor %}
                            </select> :
                            <select id="bankMinTo">
                                {% for i in range(60) %}<option value="{{ "%02d"|format(i) }}" {% if i==59 %}selected{% endif %}>{{ "%02d"|format(i) }}</option>{% endfor %}
                            </select> :
                            <select id="bankSecTo">
                                {% for i in range(60) %}<option value="{{ "%02d"|format(i) }}" {% if i==59 %}selected{% endif %}>{{ "%02d"|format(i) }}</option>{% endfor %}
                            </select>
                        </div>
                    </div>
                </div>
            </div>
            
            <div class="row">
                <div class="col">
                    <label>Регион</label>
                    <select id="region">
                        <option value="">-- Выберите регион --</option>
                        {% for code, data in regions.items() %}
                        <option value="{{ code }}">{{ data.name }}</option>
                        {% endfor %}
                    </select>
                </div>
                <div class="col">
                    <label>Маска/номер карты</label>
                    <input type="text" id="cardMask" placeholder="220220%0000">
                    <small style="color: #666;">Пример: 220220%0000</small>
                </div>
            </div>
            
            <div class="btn-group">
                <button onclick="search()">🔍 Поиск</button>
                <button onclick="clearParams()" class="danger">🗑 Очистить параметры</button>
                <button onclick="clearResults()" class="danger">🗑 Очистить результаты</button>
                <button id="exportExcelBtn" onclick="exportToExcel()" class="success" disabled>📊 Экспорт в Excel</button>
            </div>
        </div>
        
        <div id="results"></div>
    </div>
    
    <script>
        function toggleDateFrames() {
            document.getElementById('transactionFrame').style.display = 
                document.getElementById('searchByTransaction').checked ? 'block' : 'none';
            document.getElementById('bankdateFrame').style.display = 
                document.getElementById('searchByBankdate').checked ? 'block' : 'none';
        }
        
        function getDateTime(prefix, isFrom) {
            if (prefix === 'trans') {
                const date = document.getElementById('transDate' + (isFrom ? 'From' : 'To')).value;
                const hour = document.getElementById('transHour' + (isFrom ? 'From' : 'To')).value;
                const min = document.getElementById('transMin' + (isFrom ? 'From' : 'To')).value;
                const sec = document.getElementById('transSec' + (isFrom ? 'From' : 'To')).value;
                return date + ' ' + hour + ':' + min + ':' + sec;
            } else {
                const date = document.getElementById('bankDate' + (isFrom ? 'From' : 'To')).value;
                const hour = document.getElementById('bankHour' + (isFrom ? 'From' : 'To')).value;
                const min = document.getElementById('bankMin' + (isFrom ? 'From' : 'To')).value;
                const sec = document.getElementById('bankSec' + (isFrom ? 'From' : 'To')).value;
                return date + ' ' + hour + ':' + min + ':' + sec;
            }
        }
        
        function search() {
            const region = document.getElementById('region').value;
            const cardMask = document.getElementById('cardMask').value;
            
            if (!region) { alert('Выберите регион'); return; }
            if (!cardMask) { alert('Введите маску карты'); return; }
            
            const searchByTransaction = document.getElementById('searchByTransaction').checked;
            const searchByBankdate = document.getElementById('searchByBankdate').checked;
            
            if (!searchByTransaction && !searchByBankdate) {
                alert('Выберите хотя бы один тип поиска');
                return;
            }
            
            const data = {
                region_code: region,
                card_mask: cardMask,
                search_by_transaction: searchByTransaction,
                search_by_bankdate: searchByBankdate
            };
            
            if (searchByTransaction) {
                data.transaction_from = getDateTime('trans', true);
                data.transaction_to = getDateTime('trans', false);
            }
            if (searchByBankdate) {
                data.bankdate_from = getDateTime('bank', true);
                data.bankdate_to = getDateTime('bank', false);
            }
            
            document.getElementById('results').innerHTML = '<div class="loading">🔍 Поиск...</div>';
            document.getElementById('exportExcelBtn').disabled = true;
            
            fetch('/api/search', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify(data)
            })
            .then(res => res.json())
            .then(data => {
                if (data.error) {
                    document.getElementById('results').innerHTML = '<div class="error">Ошибка: ' + data.error + '</div>';
                } else if (data.results && data.results.length > 0) {
                    displayResults(data.results);
                    document.getElementById('exportExcelBtn').disabled = false;
                    window.currentResults = data.results;
                } else {
                    document.getElementById('results').innerHTML = '<div class="info">Ничего не найдено</div>';
                }
            })
            .catch(err => {
                document.getElementById('results').innerHTML = '<div class="error">Ошибка: ' + err + '</div>';
            });
        }
        
        function displayResults(results) {
            if (!results || results.length === 0) {
                document.getElementById('results').innerHTML = '<div class="info">Ничего не найдено</div>';
                return;
            }
            
            let html = '<div class="results-table"><h3>📋 Результаты поиска (найдено: ' + results.length + ')</h3>';
            html += '<table><thead><tr>';
            
            const headers = Object.keys(results[0]);
            headers.forEach(key => {
                html += '<th>' + key + '</th>';
            });
            html += '</tr></thead><tbody>';
            
            results.forEach(row => {
                html += '<tr onclick="showDetail(' + JSON.stringify(row).replace(/"/g, '&quot;') + ')">';
                headers.forEach(key => {
                    let value = row[key];
                    if (value === null || value === undefined) value = '';
                    html += '<td>' + value + '</td>';
                });
                html += '</tr>';
            });
            html += '</tbody></table></div>';
            
            document.getElementById('results').innerHTML = html;
        }
        
        function showDetail(row) {
            let details = '📋 Детальная информация:\n\n';
            Object.keys(row).forEach(key => {
                details += key + ': ' + (row[key] || 'Не указано') + '\n';
            });
            alert(details);
        }
        
        function exportToExcel() {
            if (!window.currentResults || window.currentResults.length === 0) {
                alert('Нет данных для экспорта');
                return;
            }
            
            fetch('/api/export_excel', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({results: window.currentResults})
            })
            .then(res => res.json())
            .then(data => {
                if (data.error) {
                    alert('Ошибка: ' + data.error);
                } else if (data.file_url) {
                    window.open(data.file_url, '_blank');
                }
            })
            .catch(err => alert('Ошибка: ' + err));
        }
        
        function clearParams() {
            document.getElementById('region').value = '';
            document.getElementById('cardMask').value = '';
            document.getElementById('searchByTransaction').checked = true;
            document.getElementById('searchByBankdate').checked = false;
            toggleDateFrames();
            
            const today = new Date().toISOString().split('T')[0];
            document.getElementById('transDateFrom').value = today;
            document.getElementById('transDateTo').value = today;
            document.getElementById('bankDateFrom').value = today;
            document.getElementById('bankDateTo').value = today;
        }
        
        function clearResults() {
            document.getElementById('results').innerHTML = '';
            window.currentResults = [];
            document.getElementById('exportExcelBtn').disabled = true;
        }
        
        toggleDateFrames();
    </script>
    {% endif %}
</body>
</html>
'''

@web_app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = UserManager.authenticate(username, password)
        if user:
            session['user'] = user
            return redirect(url_for('index'))
        else:
            return render_template_string(HTML_TEMPLATE, error="Неверный логин или пароль", today=datetime.now().strftime('%Y-%m-%d'), regions=REGIONS_FULL)
    
    if not session.get('user'):
        return render_template_string(HTML_TEMPLATE, today=datetime.now().strftime('%Y-%m-%d'), regions=REGIONS_FULL)
    
    return render_template_string(HTML_TEMPLATE, session=session, today=datetime.now().strftime('%Y-%m-%d'), regions=REGIONS_FULL)

@web_app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@web_app.route('/api/search', methods=['POST'])
def api_search():
    if not session.get('user'):
        return jsonify({"error": "Unauthorized"}), 401
    
    data = request.json
    region_code = data.get('region_code')
    card_mask = data.get('card_mask')
    
    region_info = REGIONS_FULL.get(region_code)
    if not region_info:
        return jsonify({"error": "Регион не найден"})
    
    conditions = []
    params = []
    
    if data.get('search_by_transaction'):
        conditions.append("transactionTime BETWEEN %s AND %s")
        params.extend([data.get('transaction_from'), data.get('transaction_to')])
    
    if data.get('search_by_bankdate'):
        conditions.append("bankdate BETWEEN %s AND %s")
        params.extend([data.get('bankdate_from'), data.get('bankdate_to')])
    
    conditions.append("sellTransportCardPAN LIKE %s")
    params.append(card_mask)
    
    where_clause = " AND ".join(conditions)
    
    query = f"""
        SELECT 
            transactionTime AS 'Дата регистрации',
            bankdate AS 'Дата списания',
            sellTransportCardPAN AS 'Номер карты',
            finalTicketPrice AS 'Стоимость',
            routeName AS 'Маршрут',
            terminalOwnerName AS 'Компания'
        FROM TerminalTransaction 
        WHERE {where_clause}
        LIMIT 1000
    """
    
    try:
        conn = pymysql.connect(
            host=region_info['host'],
            user=MYSQL_USER,
            password=MYSQL_PASSWORD,
            database=region_info['database'],
            port=MYSQL_PORT,
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor,
            connect_timeout=10
        )
        
        with conn.cursor() as cursor:
            cursor.execute(query, params)
            results = cursor.fetchall()
        
        conn.close()
        
        UserManager.log_activity(session['user']['id'], "search", f"Поиск в {region_code}, маска: {card_mask}")
        
        return jsonify({"results": results, "count": len(results)})
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@web_app.route('/api/export_excel', methods=['POST'])
def api_export_excel():
    if not session.get('user'):
        return jsonify({"error": "Unauthorized"}), 401
    
    data = request.json
    results = data.get('results', [])
    
    if not results:
        return jsonify({"error": "Нет данных"}), 400
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import os
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_{timestamp}.xlsx"
        filepath = os.path.join(os.path.dirname(__file__), filename)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Транзакции"
        
        headers = list(results[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row_num, row_data in enumerate(results, 2):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=row_data.get(header, ''))
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(filepath)
        
        UserManager.log_activity(session['user']['id'], "export_excel", f"Экспортировано {len(results)} записей")
        
        return jsonify({"file_url": f"/download/{filename}"})
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@web_app.route('/download/<filename>')
def download_file(filename):
    from flask import send_file
    import os
    filepath = os.path.join(os.path.dirname(__file__), filename)
    return send_file(filepath, as_attachment=True)

def start_web_server():
    """Запуск веб-сервера в отдельном потоке"""
    def run():
        print("🌐 Запуск полноценного веб-сервера на http://localhost:8080")
        threading.Timer(1.5, lambda: webbrowser.open('http://localhost:8080')).start()
        web_app.run(host='127.0.0.1', port=8080, debug=False, use_reloader=False)
    
    thread = threading.Thread(target=run, daemon=True)
    thread.start()
    return thread