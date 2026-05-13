# web_app.py
# Система запросов к БД - полная версия

from flask import Flask, render_template_string, request, jsonify, session, redirect, url_for
import mysql.connector
import hashlib
import pymysql
import json
import secrets
import webbrowser
import threading
import re
import time
from datetime import datetime

app = Flask(__name__)
app.secret_key = secrets.token_hex(32)

# ==================== КОНФИГУРАЦИЯ ====================
MYSQL_USER = "support"
MYSQL_PASSWORD = "vdfCD3r34$def"
MYSQL_PORT = 3306

DB_TYPES = {
    "nsi": {"name": "НСИ", "db_prefix": "nsi"},
    "sdbp": {"name": "СДБП", "db_prefix": "sdbp"}
}

# ==================== РАБОТА С ЛОКАЛЬНОЙ БД (MySQL) ====================
class Database:
    def __init__(self):
        self._init_db()
    
    def _get_conn(self):
        return mysql.connector.connect(
            host='localhost',
            user='web_user',
            password='Dkflbckfd2000',
            database='web_app_db'
        )
    
    def _init_db(self):
        conn = self._get_conn()
        cursor = conn.cursor()
        
        # Таблица пользователей
        cursor.execute('''CREATE TABLE IF NOT EXISTS users (
            id INT AUTO_INCREMENT PRIMARY KEY,
            username VARCHAR(100) UNIQUE NOT NULL,
            password_hash VARCHAR(255) NOT NULL,
            full_name VARCHAR(255) NOT NULL,
            role VARCHAR(50) DEFAULT 'user',
            is_active INT DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        
        # Таблица серверов
        cursor.execute('''CREATE TABLE IF NOT EXISTS servers (
            id INT AUTO_INCREMENT PRIMARY KEY,
            region_code VARCHAR(50) UNIQUE NOT NULL,
            region_name VARCHAR(255) NOT NULL,
            host VARCHAR(255) NOT NULL,
            database_name VARCHAR(255) NOT NULL,
            db_type VARCHAR(50) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        
        # Таблица прав доступа к серверам
        cursor.execute('''CREATE TABLE IF NOT EXISTS user_server_access (
            id INT AUTO_INCREMENT PRIMARY KEY,
            user_id INT NOT NULL,
            server_id INT NOT NULL,
            can_view INT DEFAULT 1,
            can_export INT DEFAULT 0,
            UNIQUE KEY user_server_unique (user_id, server_id)
        )''')
        
        # Таблица запросов
        cursor.execute('''CREATE TABLE IF NOT EXISTS queries (
            id INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(255) NOT NULL,
            description TEXT,
            sql_text TEXT NOT NULL,
            parameters TEXT,
            server_type VARCHAR(50),
            created_by INT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        
        # Таблица прав доступа к запросам
        cursor.execute('''CREATE TABLE IF NOT EXISTS user_query_access (
            id INT AUTO_INCREMENT PRIMARY KEY,
            user_id INT NOT NULL,
            query_id INT NOT NULL,
            can_view INT DEFAULT 1,
            UNIQUE KEY user_query_unique (user_id, query_id)
        )''')
        
        # Таблица логов
        cursor.execute('''CREATE TABLE IF NOT EXISTS logs (
            id INT AUTO_INCREMENT PRIMARY KEY,
            user_id INT,
            username VARCHAR(100),
            action VARCHAR(255),
            details TEXT,
            ip_address VARCHAR(50),
            timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        
        # Админ по умолчанию (пароль: admin123)
        admin_pass = hashlib.sha256("admin123".encode()).hexdigest()
        cursor.execute("INSERT IGNORE INTO users (id, username, password_hash, full_name, role) VALUES (1, 'admin', %s, 'Administrator', 'admin')", (admin_pass,))
        
        conn.commit()
        cursor.close()
        conn.close()
    
def execute_query(self, query, params=None, fetch_one=False, fetch_all=False):
    conn = self._get_conn()
    cursor = conn.cursor(dictionary=True)
    
    try:
        # Заменяем ? на %s для MySQL
        query = query.replace('?', '%s')
        
        if params:
            cursor.execute(query, params)
        else:
            cursor.execute(query)
        
        if fetch_one:
            result = cursor.fetchone()
            return result
        elif fetch_all:
            result = cursor.fetchall()
            return result
        else:
            conn.commit()
            return cursor.lastrowid
    finally:
        cursor.close()
        conn.close()

# ==================== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ====================
def log_action(user_id, username, action, details, ip=""):
    db.execute_query(
        "INSERT INTO logs (user_id, username, action, details, ip_address) VALUES (?, ?, ?, ?, ?)",
        (user_id, username, action, details, ip)
    )

def get_user_servers(user_id, user_role):
    if user_role == 'admin':
        servers = db.execute_query("SELECT id, region_code, region_name, host, database_name, db_type FROM servers", fetch_all=True) or []
    else:
        servers = db.execute_query("""
            SELECT s.id, s.region_code, s.region_name, s.host, s.database_name, s.db_type 
            FROM servers s
            JOIN user_server_access usa ON s.id = usa.server_id
            WHERE usa.user_id = ? AND usa.can_view = 1
        """, (user_id,), fetch_all=True) or []
    
    return [{"id": s[0], "code": s[1], "name": s[2], "host": s[3], "database": s[4], "type": s[5]} for s in servers]

def get_user_queries(user_id, user_role):
    if user_role == 'admin':
        queries = db.execute_query("SELECT id, name, description FROM queries", fetch_all=True) or []
    else:
        queries = db.execute_query("""
            SELECT q.id, q.name, q.description 
            FROM queries q
            LEFT JOIN user_query_access uqa ON q.id = uqa.query_id AND uqa.user_id = ?
            WHERE uqa.can_view = 1 OR q.created_by = ?
        """, (user_id, user_id), fetch_all=True) or []
    
    return [{"id": q[0], "name": q[1], "description": q[2]} for q in queries]

def generate_password(length=10):
    alphabet = "abcdefghijkmnopqrstuvwxyzABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    return ''.join(secrets.choice(alphabet) for _ in range(length))

def get_db_connection(server_id):
    server = db.execute_query("SELECT host, database_name FROM servers WHERE id = ?", (server_id,), fetch_one=True)
    if not server:
        return None, "Сервер не найден"
    try:
        db_conn = pymysql.connect(
            host=server[0], user=MYSQL_USER, password=MYSQL_PASSWORD,
            database=server[1], port=MYSQL_PORT, charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor, connect_timeout=10
        )
        return db_conn, None
    except Exception as e:
        return None, str(e)

# ==================== HTML ШАБЛОН ====================
MAIN_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Система запросов к БД</title>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f0f2f5; }
        .header { background: #1a1a2e; color: white; padding: 15px 30px; display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 15px; }
        .logo { font-size: 22px; font-weight: bold; }
        .user-info { display: flex; gap: 15px; align-items: center; flex-wrap: wrap; }
        .btn-header { background: rgba(255,255,255,0.15); color: white; border: none; padding: 8px 18px; border-radius: 6px; cursor: pointer; text-decoration: none; font-size: 13px; }
        .btn-header:hover { background: rgba(255,255,255,0.25); }
        .container { max-width: 1600px; margin: 20px auto; padding: 0 20px; }
        .tabs { display: flex; gap: 8px; margin-bottom: 25px; flex-wrap: wrap; border-bottom: 2px solid #ddd; padding-bottom: 10px; }
        .tab { padding: 10px 24px; background: white; border: none; border-radius: 8px 8px 0 0; cursor: pointer; font-size: 14px; font-weight: 500; color: #555; }
        .tab:hover { background: #e0e0e0; }
        .tab.active { background: #4a86e8; color: white; }
        .query-desc { background: #e8f4f8; padding: 15px 20px; border-radius: 10px; margin-bottom: 20px; border-left: 4px solid #4a86e8; }
        .params-panel { background: white; padding: 25px; border-radius: 12px; margin-bottom: 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }
        .params-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); gap: 20px; margin-bottom: 25px; }
        .param-field { display: flex; flex-direction: column; }
        .param-field label { font-size: 13px; font-weight: 600; margin-bottom: 6px; color: #555; }
        .param-field label .required { color: #dc3545; }
        .param-field input, .param-field select { padding: 10px; border: 1px solid #ddd; border-radius: 6px; font-size: 14px; }
        .region-row { display: flex; gap: 15px; align-items: flex-end; flex-wrap: wrap; margin-top: 10px; padding-top: 15px; border-top: 1px solid #eee; }
        .region-select { flex: 2; min-width: 200px; }
        .search-btn button { background: #4a86e8; color: white; border: none; padding: 10px 30px; border-radius: 6px; cursor: pointer; font-size: 15px; font-weight: bold; }
        .results { background: white; border-radius: 12px; padding: 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); overflow-x: auto; }
        .results-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 15px; flex-wrap: wrap; gap: 10px; }
        .export-btn { background: #28a745; color: white; border: none; padding: 6px 15px; border-radius: 5px; cursor: pointer; font-size: 13px; }
        .clear-btn { background: #ffc107; color: #333; border: none; padding: 6px 15px; border-radius: 5px; cursor: pointer; font-size: 13px; }
        table { width: 100%; border-collapse: collapse; font-size: 13px; }
        th, td { border: 1px solid #ddd; padding: 10px; text-align: left; }
        th { background: #4a86e8; color: white; font-weight: 600; }
        tr:hover { background: #f5f5f5; cursor: pointer; }
        tr:nth-child(even) { background: #f9f9f9; }
        .loading { text-align: center; padding: 60px; color: #999; }
        .error { color: #dc3545; padding: 15px; background: #ffe0e0; border-radius: 8px; margin: 10px 0; }
        .info { color: #856404; padding: 15px; background: #fff3cd; border-radius: 8px; margin: 10px 0; }
        .modal { display: none; position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.5); z-index: 1000; overflow: auto; }
        .modal-content { background: white; margin: 3% auto; width: 90%; max-width: 900px; border-radius: 12px; max-height: 90%; overflow-y: auto; }
        .modal-header { padding: 20px 25px; background: #1a1a2e; color: white; border-radius: 12px 12px 0 0; display: flex; justify-content: space-between; align-items: center; }
        .modal-header h2 { font-size: 20px; }
        .close { cursor: pointer; font-size: 28px; line-height: 20px; }
        .modal-body { padding: 25px; }
        .section { margin-bottom: 30px; }
        .section h3 { margin-bottom: 15px; color: #333; border-left: 3px solid #4a86e8; padding-left: 12px; }
        .card { background: #f8f9fa; padding: 15px; margin: 10px 0; border-radius: 8px; display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 10px; }
        .card-info { flex: 1; }
        .card-info strong { display: block; }
        .card-info small { color: #666; font-size: 12px; }
        .btn-icon { background: none; border: none; cursor: pointer; font-size: 18px; padding: 5px 10px; border-radius: 5px; }
        .btn-icon:hover { background: #ddd; }
        .btn-add { background: #28a745; color: white; border: none; padding: 8px 16px; border-radius: 5px; cursor: pointer; margin-top: 10px; }
        .form-group { margin-bottom: 15px; }
        .form-group label { display: block; margin-bottom: 5px; font-weight: 500; }
        .form-group input, .form-group select, .form-group textarea { width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 6px; }
        .form-row { display: flex; gap: 15px; flex-wrap: wrap; }
        .form-row .form-group { flex: 1; }
        .param-builder-row { display: flex; gap: 10px; margin-bottom: 10px; align-items: center; flex-wrap: wrap; padding: 10px; background: white; border-radius: 6px; }
        .param-builder-row input, .param-builder-row select { padding: 6px 10px; border: 1px solid #ddd; border-radius: 4px; }
        .param-builder-row .param-name { width: 150px; font-weight: bold; background: #e9ecef; }
        .password-generate { display: flex; gap: 10px; align-items: center; }
        .password-generate input { flex: 1; }
        .generate-btn { background: #6c757d; color: white; border: none; padding: 8px 15px; border-radius: 5px; cursor: pointer; }
        .access-checkboxes { display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); gap: 8px; margin-top: 10px; max-height: 200px; overflow-y: auto; padding: 10px; background: #f8f9fa; border-radius: 8px; }
        .access-checkboxes label { display: flex; align-items: center; gap: 8px; font-weight: normal; cursor: pointer; }
        @media (max-width: 768px) { .params-grid { grid-template-columns: 1fr; } .tab { padding: 6px 12px; font-size: 12px; } }
    </style>
</head>
<body>
    {% if not session.user %}
    <div style="max-width: 420px; margin: 100px auto; background: white; padding: 35px; border-radius: 16px; box-shadow: 0 10px 40px rgba(0,0,0,0.1);">
        <h2 style="margin-bottom: 25px; color: #1a1a2e;">🔐 Вход в систему</h2>
        {% if error %}<div class="error">{{ error }}</div>{% endif %}
        <form method="post">
            <div class="form-group"><label>Логин</label><input type="text" name="username" required></div>
            <div class="form-group"><label>Пароль</label><input type="password" name="password" required></div>
            <button type="submit" style="width:100%; background:#1a1a2e; color:white; border:none; padding:12px; border-radius:6px; font-size:16px; cursor:pointer;">Войти</button>
        </form>
    </div>
    {% else %}
    <div class="header">
        <div class="logo">📊 Система запросов к БД</div>
        <div class="user-info">
            <span>👋 {{ session.user.full_name }} ({{ session.user.username }})</span>
            {% if session.user.role == 'admin' %}
            <button class="btn-header" onclick="openAdminServers()">🖥 Серверы</button>
            <button class="btn-header" onclick="openAdminQueries()">📋 Запросы</button>
            <button class="btn-header" onclick="openAdminUsers()">👥 Пользователи</button>
            <button class="btn-header" onclick="openAdminLogs()">📜 Логи</button>
            {% endif %}
            <a href="/logout" class="btn-header">🚪 Выйти</a>
        </div>
    </div>
    
    <div class="container">
        <div class="tabs" id="tabs">
            {% for q in queries %}
            <button class="tab" onclick="loadQuery({{ q.id }})">{{ q.name }}</button>
            {% endfor %}
        </div>
        
        <div id="queryDescBlock" style="display: none;"><div class="query-desc"><h3 id="queryNameDisplay"></h3><p id="queryDescDisplay"></p></div></div>
        
        <div class="params-panel" id="paramsPanel" style="display: none;">
            <div id="dynamicParams" class="params-grid"></div>
            <div class="region-row">
                <div class="param-field region-select"><label>🌍 Выберите сервер (регион) <span class="required">*</span></label><select id="serverSelect"><option value="">-- Выберите сервер --</option>{% for s in servers %}<option value="{{ s.id }}">{{ s.name }} ({{ s.code }}) - {{ s.type|upper }}</option>{% endfor %}</select></div>
                <div class="search-btn"><button onclick="executeQuery()">🔍 Выполнить запрос</button></div>
            </div>
        </div>
        
        <div class="results" id="results"><div class="info">📌 Выберите вкладку с запросом</div></div>
    </div>
    
    <!-- Модальное окно: Серверы -->
    <div id="modalServers" class="modal"><div class="modal-content"><div class="modal-header"><h2>🖥 Управление серверами (регионами)</h2><span class="close" onclick="closeModal('modalServers')">&times;</span></div><div class="modal-body"><div id="serversList"></div><button class="btn-add" onclick="showServerForm()">+ Добавить сервер</button><div id="serverForm" style="display:none; margin-top:20px; padding-top:20px; border-top:1px solid #ddd;"><h3>Добавить/Редактировать сервер</h3><input type="hidden" id="serverId"><div class="form-row"><div class="form-group"><label>Тип БД *</label><select id="serverDbType"><option value="nsi">НСИ</option><option value="sdbp">СДБП</option></select></div><div class="form-group"><label>Номер региона *</label><input type="text" id="serverRegionCode" placeholder="86"></div></div><div class="form-row"><div class="form-group"><label>Название региона *</label><input type="text" id="serverRegionName" placeholder="Нижневартовск"></div><div class="form-group"><label>Хост *</label><input type="text" id="serverHost" placeholder="rds86.tkp2.prod"></div></div><div class="form-group"><label>Database *</label><input type="text" id="serverDatabase" placeholder="nsi86.prod"></div><div style="margin-top:15px;"><button class="btn-add" onclick="saveServer()">Сохранить</button><button class="btn-add" style="background:#6c757d;" onclick="hideServerForm()">Отмена</button></div></div></div></div></div>
    
    <!-- Модальное окно: Запросы -->
    <div id="modalQueries" class="modal"><div class="modal-content"><div class="modal-header"><h2>📋 Управление запросами (вкладками)</h2><span class="close" onclick="closeModal('modalQueries')">&times;</span></div><div class="modal-body"><div id="queriesList"></div><button class="btn-add" onclick="showQueryForm()">+ Добавить запрос</button><div id="queryForm" style="display:none; margin-top:20px; padding-top:20px; border-top:1px solid #ddd;"><h3>Добавить/Редактировать запрос</h3><input type="hidden" id="queryId"><div class="form-group"><label>Название вкладки *</label><input type="text" id="queryName" placeholder="Поиск по карте"></div><div class="form-group"><label>Описание</label><input type="text" id="queryDesc" placeholder="Краткое описание"></div><div class="form-group"><label>SQL запрос *</label><textarea id="querySql" rows="6" placeholder="SELECT * FROM table WHERE field LIKE :param_name"></textarea><small style="color:#666;">Параметры обозначайте двоеточием, например: :card_mask, :date_from</small></div><div id="paramsBuilderContainer" style="margin-top: 15px;"><label>Настройка параметров:</label><div id="paramsBuilderList"></div></div><div class="form-group"><label>Тип сервера (оставьте пустым для всех)</label><select id="queryServerType"><option value="">Все типы</option><option value="nsi">НСИ</option><option value="sdbp">СДБП</option></select></div><div style="margin-top:15px;"><button class="btn-add" onclick="saveQuery()">Сохранить</button><button class="btn-add" style="background:#6c757d;" onclick="hideQueryForm()">Отмена</button></div></div></div></div></div>
    
    <!-- Модальное окно: Пользователи -->
    <div id="modalUsers" class="modal"><div class="modal-content"><div class="modal-header"><h2>👥 Управление пользователями</h2><span class="close" onclick="closeModal('modalUsers')">&times;</span></div><div class="modal-body"><div id="usersList"></div><button class="btn-add" onclick="showUserForm()">+ Добавить пользователя</button><div id="userForm" style="display:none; margin-top:20px; padding-top:20px; border-top:1px solid #ddd;"><h3>Добавить/Редактировать пользователя</h3><input type="hidden" id="userId"><div class="form-group"><label>Логин *</label><input type="text" id="userUsername"></div><div class="form-group"><label>Пароль</label><div class="password-generate"><input type="text" id="userPassword" readonly><button type="button" class="generate-btn" onclick="generateUserPassword()">🎲 Сгенерировать</button></div><small>При создании пароль обязателен. При редактировании - оставьте пустым, чтобы не менять</small></div><div class="form-group"><label>Полное имя *</label><input type="text" id="userFullName"></div><div class="form-group"><label>Роль</label><select id="userRole"><option value="user">Пользователь</option><option value="admin">Администратор</option></select></div><div class="form-group"><label>Доступ к серверам:</label><div id="userServerAccess" class="access-checkboxes"></div></div><div class="form-group"><label>Доступ к запросам (вкладкам):</label><div id="userQueryAccess" class="access-checkboxes"></div></div><div style="margin-top:15px;"><button class="btn-add" onclick="saveUser()">Сохранить</button><button class="btn-add" style="background:#6c757d;" onclick="hideUserForm()">Отмена</button></div></div></div></div></div>
    
    <!-- Модальное окно: Логи -->
    <div id="modalLogs" class="modal"><div class="modal-content"><div class="modal-header"><h2>📜 Лог действий</h2><span class="close" onclick="closeModal('modalLogs')">&times;</span></div><div class="modal-body"><div id="logsList"></div></div></div></div>
    
    <!-- Модальное окно деталей -->
    <div id="detailModal" class="modal"><div class="modal-content"><div class="modal-header"><h2>📋 Детальная информация</h2><span class="close" onclick="closeDetailModal()">&times;</span></div><div class="modal-body" id="detailModalContent"></div><div style="padding:15px; text-align:center; border-top:1px solid #eee;"><button onclick="closeDetailModal()" style="background:#6c757d; color:white; border:none; padding:8px 20px; border-radius:5px; cursor:pointer;">Закрыть</button></div></div></div>
    
    <script>
        let currentQuery = null;
        let currentResults = null;
        let currentQueryId = null;
        
        function loadQuery(queryId) {
            if (currentQueryId !== queryId) {
                clearResults();
            }
            currentQueryId = queryId;
            
            fetch(`/api/query/${queryId}`).then(res => res.json()).then(data => {
                currentQuery = data;
                document.getElementById('queryDescBlock').style.display = 'block';
                document.getElementById('queryNameDisplay').innerText = data.name;
                document.getElementById('queryDescDisplay').innerText = data.description || '';
                renderParams(data.parameters);
                highlightActiveTab(queryId);
                document.getElementById('paramsPanel').style.display = 'block';
            });
        }
        
        function renderParams(params) {
            const container = document.getElementById('dynamicParams');
            if (!params || params.length === 0) {
                container.innerHTML = '<div class="info">ℹ️ Этот запрос не требует параметров</div>';
                return;
            }
            let html = '';
            params.forEach(param => {
                let inputHtml = '';
                if (param.type === 'date') {
                    inputHtml = `<input type="date" id="param_${param.name}" placeholder="${param.placeholder || ''}">`;
                } else if (param.type === 'select') {
                    inputHtml = `<select id="param_${param.name}">${param.options || ''}<option value="">-- Выберите --</option></select>`;
                } else {
                    inputHtml = `<input type="text" id="param_${param.name}" placeholder="${param.placeholder || ''}">`;
                }
                html += `<div class="param-field"><label>${param.label} ${param.required ? '<span class="required">*</span>' : ''}</label>${inputHtml}</div>`;
            });
            container.innerHTML = html;
        }
        
        function highlightActiveTab(queryId) {
            document.querySelectorAll('.tab').forEach(tab => {
                if (tab.onclick && tab.onclick.toString().includes(queryId)) tab.classList.add('active');
                else tab.classList.remove('active');
            });
        }
        
        function clearResults() {
            document.getElementById('results').innerHTML = '<div class="info">📌 Результаты очищены</div>';
            currentResults = null;
        }
        
        function executeQuery() {
            if (!currentQuery) { alert('Выберите вкладку'); return; }
            const serverId = document.getElementById('serverSelect').value;
            if (!serverId) { alert('Выберите сервер'); return; }
            const params = {};
            if (currentQuery.parameters) {
                currentQuery.parameters.forEach(param => {
                    const input = document.getElementById(`param_${param.name}`);
                    if (input && input.value) params[param.name] = input.value;
                });
            }
            document.getElementById('results').innerHTML = '<div class="loading">⏳ Выполнение запроса...</div>';
            fetch('/api/execute', {
                method: 'POST', headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({query_id: currentQuery.id, server_id: serverId, parameters: params})
            }).then(res => res.json()).then(data => {
                if (data.error) document.getElementById('results').innerHTML = '<div class="error">❌ ' + data.error + '</div>';
                else if (data.results && data.results.length > 0) { currentResults = data.results; displayResults(data.results); }
                else document.getElementById('results').innerHTML = '<div class="info">📭 Ничего не найдено</div>';
            }).catch(err => { document.getElementById('results').innerHTML = '<div class="error">Ошибка: ' + err + '</div>'; });
        }
        
        function displayResults(results) {
            if (!results || results.length === 0) {
                document.getElementById('results').innerHTML = '<div class="info">Ничего не найдено</div>';
                return;
            }
            let headers = Object.keys(results[0]);
            let html = `<div class="results-header"><div>✅ Найдено записей: ${results.length}</div><div><button class="clear-btn" onclick="clearResults()">🗑 Очистить</button><button class="export-btn" onclick="exportToExcel()" style="margin-left:10px;">📊 Экспорт в Excel</button></div></div><div style="overflow-x: auto;"><table class="results-table"><thead><tr>`;
            headers.forEach(h => html += `<th>${h}</th>`);
            html += `</thead><tbody>`;
            results.forEach((row, idx) => {
                html += `<tr onclick="showDetail(${idx})" style="cursor: pointer;">`;
                headers.forEach(h => {
                    let val = row[h];
                    if (val === null || val === undefined) val = '';
                    html += `<td>${val}</td>`;
                });
                html += '</tr>';
            });
            html += `</tbody></table></div>`;
            document.getElementById('results').innerHTML = html;
        }
        
        function showDetail(rowIndex) {
            let row = currentResults[rowIndex];
            if (!row) return;
            let detailsHtml = '<table style="width:100%; border-collapse:collapse;">';
            for (let [key, value] of Object.entries(row)) {
                let displayValue = value === null || value === undefined ? 'Не указано' : value;
                detailsHtml += `<tr style="border-bottom:1px solid #eee;"><td style="padding:10px; font-weight:bold; width:40%; background:#f8f9fa;">${key}</td><td style="padding:10px;">${displayValue}</table></tr>`;
            }
            detailsHtml += '</table>';
            document.getElementById('detailModalContent').innerHTML = detailsHtml;
            document.getElementById('detailModal').style.display = 'block';
        }
        
        function closeDetailModal() { document.getElementById('detailModal').style.display = 'none'; }
        
        function exportToExcel() {
            if (!currentResults || currentResults.length === 0) { alert('Нет данных для экспорта'); return; }
            fetch('/api/export_excel', { method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify({results: currentResults}) })
            .then(res => res.json()).then(data => { if (data.file_url) window.open(data.file_url, '_blank'); else alert('Ошибка экспорта'); });
        }
        
        // Администрирование
        function openAdminServers() { document.getElementById('modalServers').style.display = 'block'; loadServers(); }
        function openAdminQueries() { document.getElementById('modalQueries').style.display = 'block'; loadQueries(); }
        function openAdminUsers() { document.getElementById('modalUsers').style.display = 'block'; loadUsers(); }
        function openAdminLogs() { document.getElementById('modalLogs').style.display = 'block'; loadLogs(); }
        function closeModal(modalId) { document.getElementById(modalId).style.display = 'none'; }
        window.onclick = function(event) { if (event.target.classList.contains('modal')) event.target.style.display = 'none'; }
        
        function parseSqlParams() {
            let sql = document.getElementById('querySql').value;
            let paramMatches = sql.match(/:(\w+)/g);
            if (!paramMatches) { document.getElementById('paramsBuilderList').innerHTML = '<small>Параметры не найдены (используйте :имя_параметра)</small>'; return; }
            let uniqueParams = [...new Set(paramMatches.map(p => p.substring(1)))];
            let paramsJson = document.getElementById('queryParamsJson')?.value || '[]';
            let existingParams = [];
            try { existingParams = JSON.parse(paramsJson); } catch(e) { existingParams = []; }
            let html = '';
            uniqueParams.forEach(param => {
                let existing = existingParams.find(p => p.name === param) || {};
                html += `<div class="param-builder-row">
                    <input type="text" class="param-name" value="${param}" readonly style="background:#e9ecef; width:150px;">
                    <input type="text" placeholder="Название поля" id="param_label_${param}" value="${existing.label || param.replace('_', ' ').title()}" style="width:180px;">
                    <select id="param_type_${param}" style="width:100px;">
                        <option value="text" ${existing.type === 'text' ? 'selected' : ''}>Текст</option>
                        <option value="date" ${existing.type === 'date' ? 'selected' : ''}>Дата</option>
                    </select>
                    <label style="display:flex; align-items:center;"><input type="checkbox" id="param_required_${param}" ${existing.required ? 'checked' : ''}> Обязательный</label>
                    <input type="text" placeholder="Подсказка" id="param_placeholder_${param}" value="${existing.placeholder || ''}" style="width:150px;">
                </div>`;
            });
            document.getElementById('paramsBuilderList').innerHTML = html;
        }
        
        function collectParamsFromBuilder() {
            let sql = document.getElementById('querySql').value;
            let paramMatches = sql.match(/:(\w+)/g);
            if (!paramMatches) return [];
            let uniqueParams = [...new Set(paramMatches.map(p => p.substring(1)))];
            let params = [];
            uniqueParams.forEach(param => {
                params.push({
                    name: param,
                    label: document.getElementById(`param_label_${param}`)?.value || param,
                    type: document.getElementById(`param_type_${param}`)?.value || 'text',
                    required: document.getElementById(`param_required_${param}`)?.checked || false,
                    placeholder: document.getElementById(`param_placeholder_${param}`)?.value || ''
                });
            });
            return params;
        }
        
        function loadServers() {
            fetch('/api/admin/servers').then(res => res.json()).then(data => {
                let html = '';
                data.forEach(s => { html += `<div class="card"><div class="card-info"><strong>${s.region_name} (${s.region_code})</strong><small>Тип: ${s.db_type === 'nsi' ? 'НСИ' : 'СДБП'} | Хост: ${s.host} | БД: ${s.database_name}</small></div><div><button class="btn-icon" onclick="editServer(${s.id})">✏️</button><button class="btn-icon" onclick="deleteServer(${s.id})">🗑</button></div></div>`; });
                document.getElementById('serversList').innerHTML = html || '<p>Нет серверов</p>';
            });
        }
        
        function showServerForm() { document.getElementById('serverForm').style.display = 'block'; document.getElementById('serverId').value = ''; document.getElementById('serverDbType').value = 'nsi'; document.getElementById('serverRegionCode').value = ''; document.getElementById('serverRegionName').value = ''; document.getElementById('serverHost').value = ''; document.getElementById('serverDatabase').value = ''; }
        function hideServerForm() { document.getElementById('serverForm').style.display = 'none'; }
        function editServer(id) { fetch(`/api/admin/server/${id}`).then(res => res.json()).then(s => { document.getElementById('serverForm').style.display = 'block'; document.getElementById('serverId').value = s.id; document.getElementById('serverDbType').value = s.db_type; document.getElementById('serverRegionCode').value = s.region_code; document.getElementById('serverRegionName').value = s.region_name; document.getElementById('serverHost').value = s.host; document.getElementById('serverDatabase').value = s.database_name; }); }
        function saveServer() {
            let data = { id: document.getElementById('serverId').value || null, db_type: document.getElementById('serverDbType').value, region_code: document.getElementById('serverRegionCode').value, region_name: document.getElementById('serverRegionName').value, host: document.getElementById('serverHost').value, database_name: document.getElementById('serverDatabase').value };
            if (!data.region_code || !data.region_name || !data.host || !data.database_name) { alert('Заполните все поля'); return; }
            fetch('/api/admin/save_server', { method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify(data) }).then(res => res.json()).then(data => { if (data.success) { alert('Сервер сохранен'); hideServerForm(); loadServers(); } else alert('Ошибка: ' + data.error); });
        }
        function deleteServer(id) { if (confirm('Удалить сервер?')) { fetch(`/api/admin/delete_server/${id}`, {method: 'DELETE'}).then(() => loadServers()); } }
        
        function loadQueries() {
            fetch('/api/admin/queries').then(res => res.json()).then(data => {
                let html = '';
                data.forEach(q => { html += `<div class="card"><div class="card-info"><strong>${q.name}</strong><small>${q.description || 'Без описания'}</small></div><div><button class="btn-icon" onclick="editQuery(${q.id})">✏️</button><button class="btn-icon" onclick="deleteQuery(${q.id})">🗑</button></div></div>`; });
                document.getElementById('queriesList').innerHTML = html || '<p>Нет запросов</p>';
            });
        }
        
        function showQueryForm() {
            document.getElementById('queryForm').style.display = 'block';
            document.getElementById('queryId').value = '';
            document.getElementById('queryName').value = '';
            document.getElementById('queryDesc').value = '';
            document.getElementById('querySql').value = '';
            document.getElementById('queryServerType').value = '';
            document.getElementById('paramsBuilderList').innerHTML = '';
        }
        
        function hideQueryForm() { document.getElementById('queryForm').style.display = 'none'; }
        
        function editQuery(id) {
            fetch(`/api/query/${id}`).then(res => res.json()).then(q => {
                document.getElementById('queryForm').style.display = 'block';
                document.getElementById('queryId').value = q.id;
                document.getElementById('queryName').value = q.name;
                document.getElementById('queryDesc').value = q.description || '';
                document.getElementById('querySql').value = q.sql_text;
                document.getElementById('queryServerType').value = q.server_type || '';
                setTimeout(() => {
                    let paramMatches = q.sql_text.match(/:(\w+)/g);
                    if (paramMatches) {
                        let uniqueParams = [...new Set(paramMatches.map(p => p.substring(1)))];
                        let html = '';
                        uniqueParams.forEach(param => {
                            let existing = (q.parameters || []).find(p => p.name === param) || {};
                            html += `<div class="param-builder-row">
                                <input type="text" class="param-name" value="${param}" readonly style="width:150px; background:#e9ecef;">
                                <input type="text" placeholder="Название поля" id="param_label_${param}" value="${existing.label || param.replace('_', ' ').title()}" style="width:180px;">
                                <select id="param_type_${param}" style="width:100px;">
                                    <option value="text" ${existing.type === 'text' ? 'selected' : ''}>Текст</option>
                                    <option value="date" ${existing.type === 'date' ? 'selected' : ''}>Дата</option>
                                </select>
                                <label><input type="checkbox" id="param_required_${param}" ${existing.required ? 'checked' : ''}> Обязательный</label>
                                <input type="text" placeholder="Подсказка" id="param_placeholder_${param}" value="${existing.placeholder || ''}" style="width:150px;">
                            </div>`;
                        });
                        document.getElementById('paramsBuilderList').innerHTML = html;
                    } else {
                        document.getElementById('paramsBuilderList').innerHTML = '<small>Параметров не найдено (используйте :имя_параметра в SQL)</small>';
                    }
                }, 100);
            });
        }
        
        function saveQuery() {
            let params = [];
            let sql = document.getElementById('querySql').value;
            let paramMatches = sql.match(/:(\w+)/g);
            if (paramMatches) {
                let uniqueParams = [...new Set(paramMatches.map(p => p.substring(1)))];
                uniqueParams.forEach(param => {
                    params.push({
                        name: param,
                        label: document.getElementById(`param_label_${param}`)?.value || param,
                        type: document.getElementById(`param_type_${param}`)?.value || 'text',
                        required: document.getElementById(`param_required_${param}`)?.checked || false,
                        placeholder: document.getElementById(`param_placeholder_${param}`)?.value || ''
                    });
                });
            }
            let data = {
                id: document.getElementById('queryId').value || null,
                name: document.getElementById('queryName').value,
                description: document.getElementById('queryDesc').value,
                sql_text: sql,
                parameters: params,
                server_type: document.getElementById('queryServerType').value
            };
            if (!data.name || !data.sql_text) { alert('Заполните название и SQL запрос'); return; }
            fetch('/api/admin/save_query', { method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify(data) }).then(res => res.json()).then(data => { if (data.success) { alert('Запрос сохранен'); hideQueryForm(); loadQueries(); location.reload(); } else alert('Ошибка: ' + data.error); });
        }
        
        function deleteQuery(id) { if (confirm('Удалить запрос?')) { fetch(`/api/admin/delete_query/${id}`, {method: 'DELETE'}).then(() => location.reload()); } }
        
        function loadUsers() {
            fetch('/api/admin/users').then(res => res.json()).then(data => {
                let html = '';
                data.forEach(u => { html += `<div class="card"><div class="card-info"><strong>${u.username}</strong> (${u.full_name})<small>Роль: ${u.role === 'admin' ? 'Администратор' : 'Пользователь'}</small></div><div><button class="btn-icon" onclick="editUser(${u.id})">✏️</button>${u.username !== 'admin' ? `<button class="btn-icon" onclick="deleteUser(${u.id})">🗑</button>` : ''}</div></div>`; });
                document.getElementById('usersList').innerHTML = html || '<p>Нет пользователей</p>';
            });
        }
        
        function generateUserPassword() { fetch('/api/generate_password').then(res => res.json()).then(data => { document.getElementById('userPassword').value = data.password; }); }
        
        function showUserForm() {
            document.getElementById('userForm').style.display = 'block';
            document.getElementById('userId').value = '';
            document.getElementById('userUsername').value = '';
            document.getElementById('userPassword').value = '';
            document.getElementById('userFullName').value = '';
            document.getElementById('userRole').value = 'user';
            loadServerCheckboxes();
            loadQueryCheckboxes();
        }
        
        function hideUserForm() { document.getElementById('userForm').style.display = 'none'; }
        
        function loadServerCheckboxes() {
            fetch('/api/admin/servers').then(res => res.json()).then(servers => {
                let html = ''; servers.forEach(s => { html += `<label><input type="checkbox" class="serverAccess" value="${s.id}"> ${s.region_name} (${s.region_code}) - ${s.db_type === 'nsi' ? 'НСИ' : 'СДБП'}</label>`; });
                document.getElementById('userServerAccess').innerHTML = html;
            });
        }
        
        function loadQueryCheckboxes() {
            fetch('/api/admin/queries').then(res => res.json()).then(queries => {
                let html = ''; queries.forEach(q => { html += `<label><input type="checkbox" class="queryAccess" value="${q.id}"> ${q.name}</label>`; });
                document.getElementById('userQueryAccess').innerHTML = html;
            });
        }
        
        function editUser(id) {
            fetch(`/api/admin/user/${id}`).then(res => res.json()).then(u => {
                document.getElementById('userForm').style.display = 'block';
                document.getElementById('userId').value = u.id;
                document.getElementById('userUsername').value = u.username;
                document.getElementById('userPassword').value = '';
                document.getElementById('userFullName').value = u.full_name;
                document.getElementById('userRole').value = u.role;
                fetch('/api/admin/servers').then(res => res.json()).then(servers => {
                    let serverHtml = ''; servers.forEach(s => { let checked = u.server_access.includes(s.id) ? 'checked' : ''; serverHtml += `<label><input type="checkbox" class="serverAccess" value="${s.id}" ${checked}> ${s.region_name} (${s.region_code}) - ${s.db_type === 'nsi' ? 'НСИ' : 'СДБП'}</label>`; });
                    document.getElementById('userServerAccess').innerHTML = serverHtml;
                });
                fetch('/api/admin/queries').then(res => res.json()).then(queries => {
                    let queryHtml = ''; queries.forEach(q => { let checked = u.query_access.includes(q.id) ? 'checked' : ''; queryHtml += `<label><input type="checkbox" class="queryAccess" value="${q.id}" ${checked}> ${q.name}</label>`; });
                    document.getElementById('userQueryAccess').innerHTML = queryHtml;
                });
            });
        }
        
        function saveUser() {
            let serverAccess = []; document.querySelectorAll('.serverAccess:checked').forEach(cb => serverAccess.push(parseInt(cb.value)));
            let queryAccess = []; document.querySelectorAll('.queryAccess:checked').forEach(cb => queryAccess.push(parseInt(cb.value)));
            let data = { id: document.getElementById('userId').value || null, username: document.getElementById('userUsername').value, password: document.getElementById('userPassword').value, full_name: document.getElementById('userFullName').value, role: document.getElementById('userRole').value, server_access: serverAccess, query_access: queryAccess };
            if (!data.username || !data.full_name) { alert('Заполните логин и полное имя'); return; }
            fetch('/api/admin/save_user', { method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify(data) }).then(res => res.json()).then(data => { if (data.success) { alert(data.message); if (data.password_shown) alert(`Пароль для пользователя: ${data.password_shown}\\nСохраните его!`); hideUserForm(); loadUsers(); } else alert('Ошибка: ' + data.error); });
        }
        
        function deleteUser(id) { if (confirm('Удалить пользователя?')) { fetch(`/api/admin/delete_user/${id}`, {method: 'DELETE'}).then(() => loadUsers()); } }
        
        function loadLogs() {
            fetch('/api/admin/logs').then(res => res.json()).then(data => {
                let html = '<div style="overflow-x:auto;"><table><thead><tr><th>Дата</th><th>Пользователь</th><th>Действие</th><th>Детали</th></tr></thead><tbody>';
                data.forEach(l => { html += `<tr><td>${l.timestamp}</td><td>${l.username}</td><td>${l.action}</td><td>${l.details}</td></tr>`; });
                html += '</tbody></table></div>';
                document.getElementById('logsList').innerHTML = html || '<p>Нет записей</p>';
            });
        }
    </script>
    {% endif %}
</body>
</html>
'''

# ==================== API МАРШРУТЫ ====================
@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        username = request.form.get('username')
        password = hashlib.sha256(request.form.get('password', '').encode()).hexdigest()
        user = db.execute_query("SELECT id, username, full_name, role FROM users WHERE username=%s AND password_hash=%s AND is_active=1", (username, password), fetch_one=True)
        if user:
            session['user'] = {'id': user[0], 'username': user[1], 'full_name': user[2], 'role': user[3]}
            log_action(user[0], user[1], "login", "Вход в систему", request.remote_addr)
            return redirect(url_for('index'))
        return render_template_string(MAIN_TEMPLATE, error="Неверный логин или пароль", servers=[], queries=[])
    
    if not session.get('user'):
        return render_template_string(MAIN_TEMPLATE, servers=[], queries=[])
    
    servers = get_user_servers(session['user']['id'], session['user']['role'])
    queries = get_user_queries(session['user']['id'], session['user']['role'])
    
    return render_template_string(MAIN_TEMPLATE, session=session, servers=servers, queries=queries)

@app.route('/logout')
def logout():
    if session.get('user'):
        log_action(session['user']['id'], session['user']['username'], "logout", "Выход из системы", request.remote_addr)
    session.clear()
    return redirect(url_for('index'))

@app.route('/api/generate_password')
def api_generate_password():
    return jsonify({"password": generate_password()})

@app.route('/api/query/<int:query_id>')
def api_get_query(query_id):
    query = db.execute_query("SELECT id, name, description, sql_text, parameters, server_type FROM queries WHERE id=%s", (query_id,), fetch_one=True)
    if not query:
        return jsonify({"error": "Not found"}), 404
    
    sql_text = query[3]
    
    # Ищем параметры в SQL (все что с двоеточием)
    import re
    found_params = re.findall(r':(\w+)', sql_text)
    unique_params = list(dict.fromkeys(found_params))
    
    # Создаем параметры автоматически из найденных
    parameters = []
    for param in unique_params:
        parameters.append({
            "name": param,
            "label": param.replace('_', ' ').title(),
            "type": "text",
            "required": False,
            "placeholder": ""
        })
    
    # Если есть сохраненные параметры в БД - используем их настройки
    if query[4]:
        try:
            saved_params = json.loads(query[4])
            param_dict = {p['name']: p for p in parameters}
            for sp in saved_params:
                if sp['name'] in param_dict:
                    param_dict[sp['name']].update(sp)
            parameters = list(param_dict.values())
        except:
            pass
    
    return jsonify({
        "id": query[0], 
        "name": query[1], 
        "description": query[2],
        "sql_text": sql_text, 
        "parameters": parameters,
        "server_type": query[5] or ""
    })

@app.route('/api/execute', methods=['POST'])
def api_execute():
    if not session.get('user'):
        return jsonify({"error": "Unauthorized"}), 401
    
    data = request.json
    query_id = data.get('query_id')
    server_id = data.get('server_id')
    parameters = data.get('parameters', {})
    
    # Получаем базовый SQL запрос
    query = db.execute_query("SELECT name, sql_text FROM queries WHERE id=?", (query_id,), fetch_one=True)
    if not query:
        return jsonify({"error": "Query not found"}), 404
    
    base_sql = query[1]
    
    # Строим WHERE условия только для заполненных параметров
    where_conditions = []
    params_for_execute = []
    
    if parameters.get('serial'):
        where_conditions.append("t.TerminalSerialNumber LIKE %s")
        params_for_execute.append(f"%{parameters['serial']}%")
    
    if parameters.get('tuid'):
        where_conditions.append("JSON_UNQUOTE(JSON_EXTRACT(t.TerminalConfig, '$.EMV.OfflineTkpTerminalId')) LIKE %s")
        params_for_execute.append(f"%{parameters['tuid']}%")
    
    if parameters.get('bank_tid'):
        where_conditions.append("e.ID LIKE %s")
        params_for_execute.append(f"%{parameters['bank_tid']}%")
    
    # Если нет ни одного условия, возвращаем ошибку
    if not where_conditions:
        return jsonify({"error": "Заполните хотя бы одно поле для поиска"}), 400
    
    # Собираем финальный SQL
    where_clause = " AND ".join(where_conditions)
    final_sql = base_sql.replace("{where_conditions}", where_clause)
    
    print(f"DEBUG: Final SQL: {final_sql}")
    print(f"DEBUG: Params: {params_for_execute}")
    
    # Подключаемся к БД
    db_conn, error = get_db_connection(server_id)
    if error:
        return jsonify({"error": error}), 500
    
    try:
        with db_conn.cursor() as cursor:
            cursor.execute(final_sql, params_for_execute)
            results = cursor.fetchall()
        db_conn.close()
        log_action(session['user']['id'], session['user']['username'], "execute_query", f"Выполнен запрос ID={query_id}, сервер={server_id}", request.remote_addr)
        
        # Добавляем нумерацию
        for i, row in enumerate(results, 1):
            row['№'] = i
        
        return jsonify({"results": results, "count": len(results)})
    except Exception as e:
        db_conn.close()
        print(f"DEBUG: SQL Error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/export_excel', methods=['POST'])
def api_export_excel():
    if not session.get('user'):
        return jsonify({"error": "Unauthorized"}), 401
    
    data = request.json
    results = data.get('results', [])
    if not results:
        return jsonify({"error": "Нет данных"}), 400
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        import os
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_{timestamp}.xlsx"
        filepath = os.path.join(os.path.dirname(__file__), filename)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Результаты"
        headers = list(results[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row_num, row_data in enumerate(results, 2):
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num, value=str(row_data.get(header, '')))
        wb.save(filepath)
        log_action(session['user']['id'], session['user']['username'], "export_excel", f"Экспортировано {len(results)} записей", request.remote_addr)
        return jsonify({"file_url": f"/download/{filename}"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/download/<filename>')
def download_file(filename):
    from flask import send_file
    import os
    filepath = os.path.join(os.path.dirname(__file__), filename)
    return send_file(filepath, as_attachment=True)

# ==================== АДМИН API ====================
@app.route('/api/admin/servers')
def api_admin_servers():
    if not session.get('user') or session['user']['role'] != 'admin':
        return jsonify({"error": "Unauthorized"}), 401
    servers = db.execute_query("SELECT id, region_code, region_name, host, database_name, db_type FROM servers", fetch_all=True) or []
    return jsonify([{"id": s[0], "region_code": s[1], "region_name": s[2], "host": s[3], "database_name": s[4], "db_type": s[5]} for s in servers])

@app.route('/api/admin/server/<int:server_id>')
def api_admin_server(server_id):
    if not session.get('user') or session['user']['role'] != 'admin':
        return jsonify({"error": "Unauthorized"}), 401
    server = db.execute_query("SELECT id, region_code, region_name, host, database_name, db_type FROM servers WHERE id=?", (server_id,), fetch_one=True)
    if not server:
        return jsonify({"error": "Not found"}), 404
    return jsonify({"id": server[0], "region_code": server[1], "region_name": server[2], "host": server[3], "database_name": server[4], "db_type": server[5]})

@app.route('/api/admin/save_server', methods=['POST'])
def api_admin_save_server():
    if not session.get('user') or session['user']['role'] != 'admin':
        return jsonify({"error": "Unauthorized"}), 401
    data = request.json
    try:
        if data.get('id'):
            db.execute_query("UPDATE servers SET region_code=?, region_name=?, host=?, database_name=?, db_type=? WHERE id=?", (data['region_code'], data['region_name'], data['host'], data['database_name'], data['db_type'], data['id']))
            log_action(session['user']['id'], session['user']['username'], "edit_server", f"Изменен сервер {data['region_code']}", request.remote_addr)
        else:
            db.execute_query("INSERT INTO servers (region_code, region_name, host, database_name, db_type) VALUES (?, ?, ?, ?, ?)", (data['region_code'], data['region_name'], data['host'], data['database_name'], data['db_type']))
            log_action(session['user']['id'], session['user']['username'], "add_server", f"Добавлен сервер {data['region_code']}", request.remote_addr)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route('/api/admin/delete_server/<int:server_id>', methods=['DELETE'])
def api_admin_delete_server(server_id):
    if not session.get('user') or session['user']['role'] != 'admin':
        return jsonify({"error": "Unauthorized"}), 401
    db.execute_query("DELETE FROM servers WHERE id=?", (server_id,))
    db.execute_query("DELETE FROM user_server_access WHERE server_id=?", (server_id,))
    log_action(session['user']['id'], session['user']['username'], "delete_server", f"Удален сервер ID={server_id}", request.remote_addr)
    return jsonify({"success": True})

@app.route('/api/admin/queries')
def api_admin_queries():
    if not session.get('user') or session['user']['role'] != 'admin':
        return jsonify({"error": "Unauthorized"}), 401
    queries = db.execute_query("SELECT id, name, description FROM queries", fetch_all=True) or []
    return jsonify([{"id": q[0], "name": q[1], "description": q[2]} for q in queries])

@app.route('/api/admin/save_query', methods=['POST'])
def api_admin_save_query():
    if not session.get('user') or session['user']['role'] != 'admin':
        return jsonify({"error": "Unauthorized"}), 401
    data = request.json
    params_json = json.dumps(data.get('parameters', []))
    try:
        if data.get('id'):
            db.execute_query("UPDATE queries SET name=?, description=?, sql_text=?, parameters=?, server_type=? WHERE id=?", (data['name'], data.get('description', ''), data['sql_text'], params_json, data.get('server_type', ''), data['id']))
            log_action(session['user']['id'], session['user']['username'], "edit_query", f"Изменен запрос {data['name']}", request.remote_addr)
        else:
            db.execute_query("INSERT INTO queries (name, description, sql_text, parameters, server_type, created_by) VALUES (?, ?, ?, ?, ?, ?)", (data['name'], data.get('description', ''), data['sql_text'], params_json, data.get('server_type', ''), session['user']['id']))
            log_action(session['user']['id'], session['user']['username'], "add_query", f"Добавлен запрос {data['name']}", request.remote_addr)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route('/api/admin/delete_query/<int:query_id>', methods=['DELETE'])
def api_admin_delete_query(query_id):
    if not session.get('user') or session['user']['role'] != 'admin':
        return jsonify({"error": "Unauthorized"}), 401
    db.execute_query("DELETE FROM queries WHERE id=?", (query_id,))
    db.execute_query("DELETE FROM user_query_access WHERE query_id=?", (query_id,))
    log_action(session['user']['id'], session['user']['username'], "delete_query", f"Удален запрос ID={query_id}", request.remote_addr)
    return jsonify({"success": True})

@app.route('/api/admin/users')
def api_admin_users():
    if not session.get('user') or session['user']['role'] != 'admin':
        return jsonify({"error": "Unauthorized"}), 401
    users = db.execute_query("SELECT id, username, full_name, role FROM users", fetch_all=True) or []
    return jsonify([{"id": u[0], "username": u[1], "full_name": u[2], "role": u[3]} for u in users])

@app.route('/api/admin/user/<int:user_id>')
def api_admin_user(user_id):
    if not session.get('user') or session['user']['role'] != 'admin':
        return jsonify({"error": "Unauthorized"}), 401
    user = db.execute_query("SELECT id, username, full_name, role FROM users WHERE id=%s", (user_id,), fetch_one=True)
    server_access = db.execute_query("SELECT server_id FROM user_server_access WHERE user_id=?", (user_id,), fetch_all=True) or []
    query_access = db.execute_query("SELECT query_id FROM user_query_access WHERE user_id=?", (user_id,), fetch_all=True) or []
    if not user:
        return jsonify({"error": "Not found"}), 404
    return jsonify({
        "id": user[0], "username": user[1], "full_name": user[2], "role": user[3],
        "server_access": [a[0] for a in server_access],
        "query_access": [a[0] for a in query_access]
    })

@app.route('/api/admin/save_user', methods=['POST'])
def api_admin_save_user():
    if not session.get('user') or session['user']['role'] != 'admin':
        return jsonify({"error": "Unauthorized"}), 401
    
    data = request.json
    password_shown = None
    
    try:
        if data.get('id'):
            user_id = int(data['id'])
            if data.get('password') and data['password'].strip():
                password_hash = hashlib.sha256(data['password'].encode()).hexdigest()
                db.execute_query("UPDATE users SET username=?, password_hash=?, full_name=?, role=? WHERE id=?", (data['username'], password_hash, data['full_name'], data['role'], user_id))
                log_action(session['user']['id'], session['user']['username'], "edit_user", f"Изменен пользователь {data['username']} (с изменением пароля)", request.remote_addr)
            else:
                db.execute_query("UPDATE users SET username=?, full_name=?, role=? WHERE id=?", (data['username'], data['full_name'], data['role'], user_id))
                log_action(session['user']['id'], session['user']['username'], "edit_user", f"Изменен пользователь {data['username']}", request.remote_addr)
            
            db.execute_query("DELETE FROM user_server_access WHERE user_id=?", (user_id,))
            for server_id in data.get('server_access', []):
                db.execute_query("INSERT INTO user_server_access (user_id, server_id, can_view) VALUES (?, ?, 1)", (user_id, int(server_id)))
            
            db.execute_query("DELETE FROM user_query_access WHERE user_id=?", (user_id,))
            for query_id in data.get('query_access', []):
                db.execute_query("INSERT INTO user_query_access (user_id, query_id, can_view) VALUES (?, ?, 1)", (user_id, int(query_id)))
            
            message = "Пользователь обновлен"
        else:
            if not data.get('password') or not data['password'].strip():
                return jsonify({"error": "Для нового пользователя укажите пароль"}), 400
            
            existing = db.execute_query("SELECT id FROM users WHERE username=?", (data['username'],), fetch_one=True)
            if existing:
                return jsonify({"error": f"Пользователь с логином '{data['username']}' уже существует"}), 400
            
            password_hash = hashlib.sha256(data['password'].encode()).hexdigest()
            user_id = db.execute_query("INSERT INTO users (username, password_hash, full_name, role) VALUES (?, ?, ?, ?)", (data['username'], password_hash, data['full_name'], data['role']))
            
            for server_id in data.get('server_access', []):
                db.execute_query("INSERT INTO user_server_access (user_id, server_id, can_view) VALUES (?, ?, 1)", (user_id, int(server_id)))
            
            for query_id in data.get('query_access', []):
                db.execute_query("INSERT INTO user_query_access (user_id, query_id, can_view) VALUES (?, ?, 1)", (user_id, int(query_id)))
            
            password_shown = data['password']
            log_action(session['user']['id'], session['user']['username'], "add_user", f"Добавлен пользователь {data['username']}", request.remote_addr)
            message = "Пользователь создан"
        
        result = {"success": True, "message": message}
        if password_shown:
            result["password_shown"] = password_shown
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route('/api/admin/delete_user/<int:user_id>', methods=['DELETE'])
def api_admin_delete_user(user_id):
    if not session.get('user') or session['user']['role'] != 'admin':
        return jsonify({"error": "Unauthorized"}), 401
    user = db.execute_query("SELECT username FROM users WHERE id=?", (user_id,), fetch_one=True)
    if user and user[0] == 'admin':
        return jsonify({"error": "Нельзя удалить администратора"}), 400
    db.execute_query("DELETE FROM users WHERE id=?", (user_id,))
    db.execute_query("DELETE FROM user_server_access WHERE user_id=?", (user_id,))
    db.execute_query("DELETE FROM user_query_access WHERE user_id=?", (user_id,))
    log_action(session['user']['id'], session['user']['username'], "delete_user", f"Удален пользователь ID={user_id}", request.remote_addr)
    return jsonify({"success": True})

@app.route('/api/admin/logs')
def api_admin_logs():
    if not session.get('user') or session['user']['role'] != 'admin':
        return jsonify({"error": "Unauthorized"}), 401
    logs = db.execute_query("SELECT id, username, action, details, timestamp FROM logs ORDER BY id DESC LIMIT 200", fetch_all=True) or []
    return jsonify([{"id": l[0], "username": l[1], "action": l[2], "details": l[3], "timestamp": l[4]} for l in logs])

# ==================== ЗАПУСК ====================
def start_web_server():
    def run():
        print("\n" + "="*60)
        print("🚀 ВЕБ-ПРИЛОЖЕНИЕ ЗАПУЩЕНО")
        print("📱 Откройте в браузере: http://localhost:8080")
        print("👤 Логин: admin | Пароль: admin123")
        print("="*60 + "\n")
        threading.Timer(1.5, lambda: webbrowser.open('http://localhost:8080')).start()
        app.run(host='0.0.0.0', port=8080, debug=False, use_reloader=False)
    
    thread = threading.Thread(target=run, daemon=True)
    thread.start()
    return thread

if __name__ == "__main__":
    start_web_server()
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("\n👋 Остановка сервера...")
