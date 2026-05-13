import sys
import pymysql
from pymysql import Error
from datetime import datetime, timedelta
import threading
import csv
import os
from typing import Optional, Dict, List, Tuple
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app_config import REGIONS_FULL, BIN_BANKS, MYSQL_USER, MYSQL_PASSWORD, MYSQL_PORT
from widgets import ModernButton, ProgressBar, DatePicker
from transaction_detail import TransactionDetailWindow
from gui_windows import UserManagementWindow, QueryManagementWindow
from user_manager import UserManager

class TransactionSearchApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Поиск транзакций ver. 2.1.6")
        self.root.geometry("1420x807")
        self.root.configure(bg='white')
        
        # Переменные для хранения данных
        self.search_thread = None
        self.stop_search_flag = False
        self.connection = None
        self.current_results = []
        self.progress_bar = None
        self.status_label = None
        self.row_count = 0
        
        # Переменные для чекбоксов
        self.search_by_transaction = tk.BooleanVar(value=True)
        self.search_by_bankdate = tk.BooleanVar(value=False)
        
        # Создаем интерфейс
        self.setup_ui()
        
        self.log_message("Приложение запущено")
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def get_route_tariffs(self, route_name):
        """
        Получение информации о тарифах для конкретного маршрута по его названию
        Работает ТОЛЬКО для 54 региона (Новосибирск)
        """
        connection = None
        try:
            # Данные для подключения к 54 региону
            region_54 = REGIONS_FULL["54"]
            
            connection = pymysql.connect(
                host=region_54["host"],
                user=MYSQL_USER,
                password=MYSQL_PASSWORD,
                database=region_54["database"],
                port=MYSQL_PORT,
                charset='utf8mb4',
                cursorclass=pymysql.cursors.DictCursor,
                connect_timeout=5
            )
            
            with connection.cursor() as cursor:
                # Находим маршрут по названию (ищем по частичному совпадению)
                find_route_query = """
                SELECT 
                    r.route_id, 
                    r.organization_id, 
                    r.number, 
                    r.name,
                    o.title_small as organization_name,
                    o.partner_id
                FROM route r
                LEFT JOIN organization o ON o.organization_id = r.organization_id
                WHERE r.name LIKE %s
                  AND r.is_archive = 0 
                  AND r.is_deleted = 0
                LIMIT 1
                """
                
                # Ищем по части названия маршрута
                search_pattern = f"%{route_name}%"
                cursor.execute(find_route_query, (search_pattern,))
                route = cursor.fetchone()
                
                if not route:
                    # Если не нашли по названию, пробуем найти по номеру маршрута
                    find_by_number_query = """
                    SELECT 
                        r.route_id, 
                        r.organization_id, 
                        r.number, 
                        r.name,
                        o.title_small as organization_name,
                        o.partner_id
                    FROM route r
                    LEFT JOIN organization o ON o.organization_id = r.organization_id
                    WHERE r.number LIKE %s
                      AND r.is_archive = 0 
                      AND r.is_deleted = 0
                    LIMIT 1
                    """
                    cursor.execute(find_by_number_query, (search_pattern,))
                    route = cursor.fetchone()
                    
                if not route:
                    return None
                
                # Получаем информацию о тарифах для найденного маршрута
                tariffs_query = """
                SELECT 
                    eob.WriteOffId,
                    eob.IsActive as org_active,
                    eob.`Default` as is_default,
                    COALESCE(eorb.Status, 0) as route_status,
                    CASE 
                        WHEN eob.IsActive = 1 
                             AND (
                                 (eob.`Default` = 1 AND COALESCE(eorb.Status, 0) IN (1,2))
                                 OR (eob.`Default` = 0 AND COALESCE(eorb.Status, 0) = 1)
                             )
                        THEN 1 
                        ELSE 0 
                    END as is_active
                FROM EMVWriteOffOrganizationBinding eob
                LEFT JOIN EMVWriteOffRouteBinding eorb 
                    ON eorb.RouteId = %s 
                    AND eorb.WriteOffId = eob.WriteOffId
                WHERE eob.OrganizationId = %s
                ORDER BY eob.WriteOffId
                """
                
                cursor.execute(tariffs_query, (route['route_id'], route['organization_id']))
                tariffs_data = cursor.fetchall()
                
                # Словарь с названиями тарифов
                tariff_names = {
                    1: 'Безлимитный',
                    4: 'КЖ Студент. Безлимитный',
                    7: 'МПК Социальная карта. 30 поездок',
                    10: 'МПК Дисконт',
                    13: 'МПК Карта Студента',
                    16: 'МПК Карта Школьника',
                    19: 'МПК Социальная карта. Безлимит',
                    22: 'Скидка 50%',
                    25: 'КЖ 30 поездок',
                    28: 'КЖ Безлими',
                    31: 'КЖ Студент',
                    34: 'КЖ Школьник',
                    35: 'Коллегия Минцифры НСО',
                    38: 'Тест',
                    39: 'Сопровождающий ребёнка инвалида',
                    42: 'КЖ Безлимит. Родитель-опекун',
                    45: 'КЖ Безлимит. Детский',
                    48: 'Выпуск карты студента',
                    114: 'ЕТК',
                    117: 'ЕТК-безлимит',
                    118: 'Единый проездной (скидка 1 рубль)',
                    119: 'Единый проездной (безлимит)',
                    120: 'МПК Социальная карта. Безлимит Многодетные',
                    121: 'МПК Карта Студента. Безлимитный',
                    122: 'Эл.кошелек для корпоративных ТК',
                    126: 'КЖ 30 поездок_скидка',
                    127: 'КЖ безлимит_скидка',
                    128: 'КЖ Студент. Безлимитный_скидка',
                    129: 'Единый проездной (безлимит)_скидка'
                }
                
                # Формируем результат
                result = {
                    'route_info': {
                        'number': route['number'] if route['number'] else 'Не указан',
                        'name': route['name'] if route['name'] else 'Не указано',
                        'organization': route['organization_name'] if route['organization_name'] else 'Не указано'
                    },
                    'tariffs': []
                }
                
                for tariff in tariffs_data:
                    if tariff['WriteOffId'] in tariff_names:
                        tariff_info = {
                            'name': tariff_names[tariff['WriteOffId']],
                            'is_active': bool(tariff['is_active']),
                            'org_active': bool(tariff['org_active']),
                            'route_status': tariff['route_status']
                        }
                        result['tariffs'].append(tariff_info)
                
                return result
                
        except Exception as e:
            print(f"Ошибка при получении информации о тарифах: {e}")
            return None
        finally:
            if connection:
                connection.close()
    
    def setup_ui(self):
        """Создание полностью белого интерфейса"""
        # Основной контейнер с прокруткой для всего окна
        main_canvas = tk.Canvas(self.root, bg='white', highlightthickness=0)
        main_scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=main_canvas.yview)
        main_scrollable_frame = ttk.Frame(main_canvas)
        
        main_scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )
        
        main_canvas.create_window((0, 0), window=main_scrollable_frame, anchor="nw")
        main_canvas.configure(yscrollcommand=main_scrollbar.set)
        
        # Добавляем скроллинг колесиком мышки для главного окна
        def _on_main_mousewheel(event):
            main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        main_canvas.bind_all("<MouseWheel>", _on_main_mousewheel)
        
        # Отвязываем событие при покидании окна
        def _unbind_main_mousewheel(event):
            main_canvas.unbind_all("<MouseWheel>")
        
        def _bind_main_mousewheel(event):
            main_canvas.bind_all("<MouseWheel>", _on_main_mousewheel)
        
        main_canvas.bind("<Enter>", _bind_main_mousewheel)
        main_canvas.bind("<Leave>", _unbind_main_mousewheel)
        
        main_canvas.pack(side="left", fill="both", expand=True)
        main_scrollbar.pack(side="right", fill="y")
        
        # Верхняя панель
        top_frame = ttk.Frame(main_scrollable_frame, padding="15")
        top_frame.pack(fill="x")
        
        # Заголовок
        title_label = ttk.Label(top_frame, text="Помощник в поиске транзакций", 
                                font=("Segoe UI", 16, "bold"))
        title_label.pack(anchor="center", pady=(0, 15))
        
        # Выбор региона - ПО ЦЕНТРУ
        region_frame = ttk.Frame(top_frame)
        region_frame.pack(fill="x", pady=(0, 10))
        
        # Центрируем регион
        region_center = ttk.Frame(region_frame)
        region_center.pack(anchor="center")
        
        ttk.Label(region_center, text="Регион:", font=("Segoe UI", 10)).pack(side="left", padx=(0, 10))
        
        # Сортируем регионы
        region_values = []
        numeric_regions = []
        non_numeric_regions = []
        
        for code, data in REGIONS_FULL.items():
            if code.replace('.', '').isdigit():
                numeric_code = int(code) if code.isdigit() else float(code)
                numeric_regions.append((numeric_code, code, data['name']))
            else:
                non_numeric_regions.append((code, data['name']))
        
        numeric_regions.sort(key=lambda x: x[0])
        non_numeric_regions.sort(key=lambda x: x[0])
        
        for numeric_code, code, name in numeric_regions:
            if str(code).isdigit() and len(str(code)) < 2:
                display_code = str(code).zfill(2)
            else:
                display_code = str(code)
            region_values.append(f"{display_code} - {name}")
        
        for code, name in non_numeric_regions:
            region_values.append(f"{code} - {name}")
        
        self.region_var = tk.StringVar()
        self.region_combo = ttk.Combobox(region_center, textvariable=self.region_var,
                                        values=region_values,
                                        state="readonly", width=50)
        self.region_combo.pack(side="left")
        
        # Чекбоксы для выбора типа поиска - ПО ЦЕНТРУ
        checkbox_frame = ttk.Frame(top_frame)
        checkbox_frame.pack(fill="x", pady=(10, 10))
        
        checkbox_center = ttk.Frame(checkbox_frame)
        checkbox_center.pack(anchor="center")
        
        self.search_by_transaction = tk.BooleanVar(value=True)
        self.search_by_bankdate = tk.BooleanVar(value=False)
        
        ttk.Checkbutton(checkbox_center, text="Поиск по дате регистрации",
                       variable=self.search_by_transaction,
                       command=self.on_transaction_check).pack(side="left", padx=(0, 20))
        
        ttk.Checkbutton(checkbox_center, text="Поиск по дате списания",
                       variable=self.search_by_bankdate,
                       command=self.on_bankdate_check).pack(side="left")
        
        # Дата и время регистрации на терминале
        self.terminal_frame = ttk.LabelFrame(top_frame, text="Дата и время регистрации на терминале", padding="12")
        self.terminal_frame.pack(fill="x", pady=(0, 10))
        
        # Контейнер для двух колонок
        terminal_columns = ttk.Frame(self.terminal_frame)
        terminal_columns.pack(fill="x", expand=True)
        
        # Левая колонка - Дата и время с
        left_column = ttk.Frame(terminal_columns)
        left_column.pack(side="left", expand=True, fill="x", padx=(20, 10))
        
        from_label = ttk.Label(left_column, text="Дата и время с:", font=("Segoe UI", 9, "bold"))
        from_label.pack(anchor="w")
        
        # Дата С
        self.date_from_picker = DatePicker(left_column)
        self.date_from_picker.pack(anchor="w", pady=(5, 0))
        
        # Время С
        time_from_frame = ttk.Frame(left_column)
        time_from_frame.pack(anchor="w", pady=(5, 0))
        
        hours = [f"{i:02d}" for i in range(24)]
        minutes = [f"{i:02d}" for i in range(60)]
        
        self.hour_from_combo = ttk.Combobox(time_from_frame, values=hours, 
                                           width=3, state="readonly")
        self.hour_from_combo.set("00")
        self.hour_from_combo.pack(side="left")
        
        ttk.Label(time_from_frame, text=":").pack(side="left", padx=2)
        
        self.minute_from_combo = ttk.Combobox(time_from_frame, values=minutes, 
                                             width=3, state="readonly")
        self.minute_from_combo.set("00")
        self.minute_from_combo.pack(side="left")
        
        ttk.Label(time_from_frame, text=":").pack(side="left", padx=2)
        
        self.second_from_combo = ttk.Combobox(time_from_frame, values=minutes, 
                                             width=3, state="readonly")
        self.second_from_combo.set("00")
        self.second_from_combo.pack(side="left")
        
        # Правая колонка - Дата и время по
        right_column = ttk.Frame(terminal_columns)
        right_column.pack(side="right", expand=True, fill="x", padx=(10, 20))
        
        to_label = ttk.Label(right_column, text="Дата и время по:", font=("Segoe UI", 9, "bold"))
        to_label.pack(anchor="w")
        
        # Дата ПО
        self.date_to_picker = DatePicker(right_column)
        self.date_to_picker.pack(anchor="w", pady=(5, 0))
        
        # Время ПО
        time_to_frame = ttk.Frame(right_column)
        time_to_frame.pack(anchor="w", pady=(5, 0))
        
        self.hour_to_combo = ttk.Combobox(time_to_frame, values=hours, 
                                         width=3, state="readonly")
        self.hour_to_combo.set("23")
        self.hour_to_combo.pack(side="left")
        
        ttk.Label(time_to_frame, text=":").pack(side="left", padx=2)
        
        self.minute_to_combo = ttk.Combobox(time_to_frame, values=minutes, 
                                           width=3, state="readonly")
        self.minute_to_combo.set("59")
        self.minute_to_combo.pack(side="left")
        
        ttk.Label(time_to_frame, text=":").pack(side="left", padx=2)
        
        self.second_to_combo = ttk.Combobox(time_to_frame, values=minutes, 
                                           width=3, state="readonly")
        self.second_to_combo.set("59")
        self.second_to_combo.pack(side="left")
        
        # Дата и время списания
        self.bankdate_frame = ttk.LabelFrame(top_frame, text="Дата и время списания", padding="12")
        
        # Контейнер для двух колонок
        bank_columns = ttk.Frame(self.bankdate_frame)
        bank_columns.pack(fill="x", expand=True)
        
        # Левая колонка - Дата и время с
        bank_left = ttk.Frame(bank_columns)
        bank_left.pack(side="left", expand=True, fill="x", padx=(20, 10))
        
        bank_from_label = ttk.Label(bank_left, text="Дата и время с:", font=("Segoe UI", 9, "bold"))
        bank_from_label.pack(anchor="w")
        
        # Дата С для списания
        self.bankdate_from_picker = DatePicker(bank_left)
        self.bankdate_from_picker.pack(anchor="w", pady=(5, 0))
        
        # Время С для списания
        bank_time_from_frame = ttk.Frame(bank_left)
        bank_time_from_frame.pack(anchor="w", pady=(5, 0))
        
        self.bank_hour_from_combo = ttk.Combobox(bank_time_from_frame, values=hours, 
                                                width=3, state="readonly")
        self.bank_hour_from_combo.set("00")
        self.bank_hour_from_combo.pack(side="left")
        
        ttk.Label(bank_time_from_frame, text=":").pack(side="left", padx=2)
        
        self.bank_minute_from_combo = ttk.Combobox(bank_time_from_frame, values=minutes, 
                                                  width=3, state="readonly")
        self.bank_minute_from_combo.set("00")
        self.bank_minute_from_combo.pack(side="left")
        
        ttk.Label(bank_time_from_frame, text=":").pack(side="left", padx=2)
        
        self.bank_second_from_combo = ttk.Combobox(bank_time_from_frame, values=minutes, 
                                                  width=3, state="readonly")
        self.bank_second_from_combo.set("00")
        self.bank_second_from_combo.pack(side="left")
        
        # Правая колонка - Дата и время по
        bank_right = ttk.Frame(bank_columns)
        bank_right.pack(side="right", expand=True, fill="x", padx=(10, 20))
        
        bank_to_label = ttk.Label(bank_right, text="Дата и время по:", font=("Segoe UI", 9, "bold"))
        bank_to_label.pack(anchor="w")
        
        # Дата ПО для списания
        self.bankdate_to_picker = DatePicker(bank_right)
        self.bankdate_to_picker.pack(anchor="w", pady=(5, 0))
        
        # Время ПО для списания
        bank_time_to_frame = ttk.Frame(bank_right)
        bank_time_to_frame.pack(anchor="w", pady=(5, 0))
        
        self.bank_hour_to_combo = ttk.Combobox(bank_time_to_frame, values=hours, 
                                              width=3, state="readonly")
        self.bank_hour_to_combo.set("23")
        self.bank_hour_to_combo.pack(side="left")
        
        ttk.Label(bank_time_to_frame, text=":").pack(side="left", padx=2)
        
        self.bank_minute_to_combo = ttk.Combobox(bank_time_to_frame, values=minutes, 
                                                width=3, state="readonly")
        self.bank_minute_to_combo.set("59")
        self.bank_minute_to_combo.pack(side="left")
        
        ttk.Label(bank_time_to_frame, text=":").pack(side="left", padx=2)
        
        self.bank_second_to_combo = ttk.Combobox(bank_time_to_frame, values=minutes, 
                                                width=3, state="readonly")
        self.bank_second_to_combo.set("59")
        self.bank_second_to_combo.pack(side="left")
        
        # Изначально показываем только дату транзакции
        self.bankdate_frame.pack_forget()
        
        # Маска карты - ПО ЦЕНТРУ
        card_frame = ttk.Frame(top_frame)
        card_frame.pack(fill="x", pady=(0, 10))
        
        card_center = ttk.Frame(card_frame)
        card_center.pack(anchor="center")
        
        ttk.Label(card_center, text="Маска/номер карты:", font=("Segoe UI", 10, "bold")).pack(side="left", padx=(0, 10))
        
        self.card_mask_var = tk.StringVar()
        card_mask_entry = ttk.Entry(card_center, textvariable=self.card_mask_var, width=40)
        card_mask_entry.pack(side="left")
        card_mask_entry.bind('<Return>', self.on_enter_pressed)
        
        ttk.Label(card_center, text="Пример: 220220%0000", 
                 font=("Segoe UI", 8), foreground="gray").pack(side="left", padx=(10, 0))
        
        # Анимированная строка загрузки - ПО ЦЕНТРУ
        self.progress_container = ttk.Frame(top_frame)
        self.progress_container.pack(fill="x", pady=5)
        
        progress_center = ttk.Frame(self.progress_container)
        progress_center.pack(anchor="center")
        
        self.progress_bar = ProgressBar(progress_center, width=350, height=22)
        self.progress_bar.grid(row=0, column=0, sticky=tk.W)
        
        # Статусная надпись под прогресс-баром
        self.status_label = tk.Label(self.progress_container, text="Готов к работе", 
                                     font=('Segoe UI', 9), fg='#333333', bg='white')
        self.status_label.pack(pady=(5, 0))
        
        # Кнопки управления - ПО ЦЕНТРУ
        buttons_frame = ttk.Frame(top_frame)
        buttons_frame.pack(pady=12)
        
        buttons_center = ttk.Frame(buttons_frame)
        buttons_center.pack(anchor="center")
        
        self.search_btn = ModernButton(buttons_center, text="Поиск", 
                                      command=self.start_search,
                                      bg_color="#4a86e8", hover_color="#3a76d8",
                                      width=120, height=38)
        self.search_btn.pack(side="left", padx=5)
        
        self.stop_btn = ModernButton(buttons_center, text="Остановить", 
                                    command=self.stop_search,
                                    bg_color="#dc3545", hover_color="#c82333",
                                    width=120, height=38, 
                                    state='disabled')
        self.stop_btn.pack(side="left", padx=5)
        
        self.clear_params_btn = ModernButton(buttons_center, text="Очистить параметры", 
                                            command=self.clear_search_params,
                                            bg_color="#6c757d", hover_color="#5a6268",
                                            width=150, height=38)
        self.clear_params_btn.pack(side="left", padx=5)
        
        self.clear_results_btn = ModernButton(buttons_center, text="Очистить результаты", 
                                             command=self.clear_results,
                                             bg_color="#6c757d", hover_color="#5a6268",
                                             width=150, height=38)
        self.clear_results_btn.pack(side="left", padx=5)
        
        self.export_excel_btn = ModernButton(buttons_center, text="Экспорт в Excel", 
                                            command=self.export_to_excel,
                                            bg_color="#28a745", hover_color="#218838",
                                            width=140, height=38, 
                                            state='disabled')
        self.export_excel_btn.pack(side="left", padx=5)
        
        # Кнопка экспорта всех транзакций в один PDF
        self.export_all_pdf_btn = ModernButton(buttons_center, text="📄 Все в PDF", 
                                              command=self.export_all_to_pdf,
                                              bg_color="#dc3545", hover_color="#c82333",
                                              width=120, height=38, 
                                              state='disabled')
        self.export_all_pdf_btn.pack(side="left", padx=5)
        
        # Область результатов
        results_frame = ttk.LabelFrame(main_scrollable_frame, text="Результаты поиска (двойной клик для деталей)", padding="12")
        results_frame.pack(fill="both", expand=True, padx=15, pady=(0, 10))
        
        self.create_results_treeview(results_frame)
    
    def get_check_url(self):
        """Получение URL для проверки билета на основе выбранного региона"""
        region_text = self.region_var.get()
        if not region_text:
            return "https://qr.sbertroika.ru/cheques"
        
        region_code = region_text.split(' - ')[0]
        
        # Для Луганска
        if region_code == "t2":
            return "https://t2.qr.sbertroika.ru/cheques"
        
        # Для регионов с дополнительными цифрами (292, 702, 742, 272)
        if len(region_code) > 2 and region_code.isdigit():
            base_code = region_code[:2]
        else:
            base_code = region_code
        
        return f"https://{base_code}.qr.sbertroika.ru/cheques"
    
    def on_transaction_check(self):
        """Обработка выбора чекбокса поиска по дате регистрации"""
        if self.search_by_transaction.get():
            self.search_by_bankdate.set(False)
        self.toggle_date_frames()
    
    def on_bankdate_check(self):
        """Обработка выбора чекбокса поиска по дате списания"""
        if self.search_by_bankdate.get():
            self.search_by_transaction.set(False)
        self.toggle_date_frames()
    
    def toggle_date_frames(self):
        """Показ/скрытие блоков с датами в зависимости от чекбоксов"""
        self.terminal_frame.pack_forget()
        self.bankdate_frame.pack_forget()
        
        if self.search_by_transaction.get():
            self.terminal_frame.pack(fill="x", pady=(0, 10))
        
        if self.search_by_bankdate.get():
            self.bankdate_frame.pack(fill="x", pady=(0, 10))
    
    def create_results_treeview(self, parent):
        """Создание Treeview для отображения результатов (без серии и номера билета)"""
        tree_frame = ttk.Frame(parent)
        tree_frame.pack(fill="both", expand=True)
        
        # Убраны только серия и номер билета (последние 2 столбца)
        columns = (
            'Дата регистрации на терминале',
            'Дата списания', 
            'Номер карты',
            'Стоимость проезда',
            'Стоимость билета',
            'Скидка',
            'С/Н терминала',
            'ГРЗ',
            'Маршрут',
            'Компания'
        )
        
        self.results_tree = ttk.Treeview(tree_frame, columns=columns, show='headings',
                                        height=15)
        
        # Настройка столбцов
        column_widths = [150, 150, 150, 100, 100, 80, 120, 100, 200, 200]
        for col, width in zip(columns, column_widths):
            self.results_tree.heading(col, text=col)
            self.results_tree.column(col, width=width, minwidth=width, anchor='center')
        
        self.results_tree.bind('<Double-1>', self.on_item_double_click)
        
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.results_tree.yview)
        self.results_tree.configure(yscrollcommand=scrollbar.set)
        
        h_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.results_tree.xview)
        self.results_tree.configure(xscrollcommand=h_scrollbar.set)
        
        self.results_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")
    
    def on_enter_pressed(self, event):
        self.start_search()
    
    def on_item_double_click(self, event):
        selection = self.results_tree.selection()
        if not selection:
            return
        
        item = self.results_tree.item(selection[0])
        values = item['values']
        
        if len(values) >= 10:
            # Ищем полную транзакцию в self.current_results по уникальным полям
            for trans in self.current_results:
                if (str(trans.get('Дата регистрации на терминале', '')) == str(values[0]) and
                    str(trans.get('Номер карты', '')) == str(values[2])):
                    transaction_data = {
                        'Дата регистрации на терминале': trans.get('Дата регистрации на терминале'),
                        'Дата списания': trans.get('Дата списания'),
                        'Номер карты': trans.get('Номер карты'),
                        'Стоимость проезда': trans.get('Стоимость проезда'),
                        'Стоимость билета': trans.get('Стоимость билета'),
                        'Скидка': trans.get('Скидка'),
                        'С/Н терминала': trans.get('С/Н терминала'),
                        'ГРЗ': trans.get('ГРЗ'),
                        'Маршрут': trans.get('Маршрут'),
                        'Компания': trans.get('Компания'),
                        'Серия билета': trans.get('Серия билета'),
                        'Номер билета': trans.get('Номер билета'),
                    }
                    TransactionDetailWindow(self.root, transaction_data, self)
                    break
    
    def get_datetime_from_ui(self):
        date_from = self.date_from_picker.get_date()
        date_to = self.date_to_picker.get_date()
        
        time_from = f"{self.hour_from_combo.get()}:{self.minute_from_combo.get()}:{self.second_from_combo.get()}"
        time_to = f"{self.hour_to_combo.get()}:{self.minute_to_combo.get()}:{self.second_to_combo.get()}"
        
        return (f"{date_from.strftime('%Y-%m-%d')} {time_from}",
                f"{date_to.strftime('%Y-%m-%d')} {time_to}")
    
    def get_bankdatetime_from_ui(self):
        date_from = self.bankdate_from_picker.get_date()
        date_to = self.bankdate_to_picker.get_date()
        
        time_from = f"{self.bank_hour_from_combo.get()}:{self.bank_minute_from_combo.get()}:{self.bank_second_from_combo.get()}"
        time_to = f"{self.bank_hour_to_combo.get()}:{self.bank_minute_to_combo.get()}:{self.bank_second_to_combo.get()}"
        
        return (f"{date_from.strftime('%Y-%m-%d')} {time_from}",
                f"{date_to.strftime('%Y-%m-%d')} {time_to}")
    
    def log_message(self, message: str):
        self.status_label.config(text=message)
        self.root.update_idletasks()
    
    def validate_inputs(self) -> Tuple[bool, str]:
        region_text = self.region_var.get()
        if not region_text:
            return False, "Не указан регион"
        
        region_code = region_text.split(' - ')[0]
        if region_code.isdigit() and len(region_code) == 1:
            region_code = region_code.zfill(2)
        
        if region_code not in REGIONS_FULL:
            return False, "Выбран некорректный регион"
        
        if not self.search_by_transaction.get() and not self.search_by_bankdate.get():
            return False, "Выберите хотя бы один тип поиска (по дате транзакции или по дате списания)"
        
        if self.search_by_transaction.get():
            try:
                date_from_str, date_to_str = self.get_datetime_from_ui()
                
                dt_from = datetime.strptime(date_from_str, '%Y-%m-%d %H:%M:%S')
                dt_to = datetime.strptime(date_to_str, '%Y-%m-%d %H:%M:%S')
                
                if dt_from > dt_to:
                    return False, "Дата регистрации 'С' не может быть позже даты 'По'"
            except ValueError as e:
                return False, f"Неверный формат даты/времени регистрации: {str(e)}"
        
        if self.search_by_bankdate.get():
            try:
                bank_from_str, bank_to_str = self.get_bankdatetime_from_ui()
                
                bank_dt_from = datetime.strptime(bank_from_str, '%Y-%m-%d %H:%M:%S')
                bank_dt_to = datetime.strptime(bank_to_str, '%Y-%m-%d %H:%M:%S')
                
                if bank_dt_from > bank_dt_to:
                    return False, "Дата списания 'С' не может быть позже даты 'По'"
            except ValueError as e:
                return False, f"Неверный формат даты/времени списания: {str(e)}"
        
        card_mask = self.card_mask_var.get().strip()
        if not card_mask:
            return False, "Не указана маска карты"
        
        return True, ""
    
    def get_region_info(self) -> Optional[Dict]:
        region_text = self.region_var.get()
        if not region_text:
            return None
        
        region_code = region_text.split(' - ')[0]
        if region_code.isdigit() and len(region_code) == 1:
            region_code = region_code.zfill(2)
        
        return REGIONS_FULL.get(region_code)
    
    def create_connection(self, region_info: Dict) -> Optional[pymysql.connections.Connection]:
        try:
            self.log_message(f"Подключение к БД {region_info['name']}...")
            
            connection = pymysql.connect(
                host=region_info['host'],
                user=MYSQL_USER,
                password=MYSQL_PASSWORD,
                database=region_info['database'],
                port=MYSQL_PORT,
                charset='utf8mb4',
                cursorclass=pymysql.cursors.DictCursor,
                connect_timeout=10
            )
            
            self.log_message(f"Успешное подключение к БД {region_info['database']}")
            return connection
                
        except Error as e:
            self.log_message(f"Ошибка подключения к БД: {str(e)}")
            messagebox.showerror("Ошибка подключения", 
                               f"Не удалось подключиться к БД:\n{str(e)}")
            return None
    
    def build_query(self) -> Tuple[str, List]:
        conditions = []
        params = []
        
        if self.search_by_transaction.get():
            conditions.append("transactionTime BETWEEN %s AND %s")
            date_from_str, date_to_str = self.get_datetime_from_ui()
            params.extend([date_from_str, date_to_str])
        
        if self.search_by_bankdate.get():
            conditions.append("bankdate BETWEEN %s AND %s")
            bank_from_str, bank_to_str = self.get_bankdatetime_from_ui()
            params.extend([bank_from_str, bank_to_str])
        
        conditions.append("sellTransportCardPAN LIKE %s")
        params.append(self.card_mask_var.get())
        
        where_clause = " AND ".join(conditions)
        
        query = f"""
        SELECT 
            transactionTime AS 'Дата регистрации на терминале',
            bankdate AS 'Дата списания',
            sellTransportCardPAN AS 'Номер карты',
            finalTicketPrice AS 'Стоимость проезда',
            baseTicketPrice AS 'Стоимость билета',
            sumForDiscount AS 'Скидка',
            terminalSerialNumber AS 'С/Н терминала',
            busNumber AS 'ГРЗ',
            routeName AS 'Маршрут',
            terminalOwnerName AS 'Компания',
            ticketSeries AS 'Серия билета',
            ticketNumber AS 'Номер билета'
        FROM TerminalTransaction 
        WHERE {where_clause}
        """
        
        return query, params
    
    def execute_search(self):
        try:
            self.current_results = []
            self.row_count = 0
            
            region_info = self.get_region_info()
            
            self.connection = self.create_connection(region_info)
            if not self.connection:
                self.root.after(0, self.search_complete)
                return
            
            with self.connection.cursor() as cursor:
                self.log_message("Выполнение запроса...")
                
                query, params = self.build_query()
                cursor.execute(query, params)
                
                results = []
                row_count = 0
                
                while True:
                    if self.stop_search_flag:
                        self.log_message("Поиск остановлен пользователем")
                        break
                    
                    batch = cursor.fetchmany(100)
                    if not batch:
                        break
                    
                    results.extend(batch)
                    self.current_results.extend(batch)
                    row_count += len(batch)
                    self.row_count = row_count
                    
                    progress = min(row_count // 10, 95)
                    self.progress_bar.set_progress(progress)
                    
                    self.log_message(f"Обработано: {row_count} записей")
                    
                    if row_count % 500 == 0:
                        self.root.after(0, lambda b=batch: self.display_batch_results(b))
                
                if results and not self.stop_search_flag:
                    self.root.after(0, lambda: self.display_results(results))
                
                self.log_message(f"Поиск завершен. Найдено {row_count} записей")
            
        except Error as e:
            self.log_message(f"Ошибка выполнения запроса: {str(e)}")
            messagebox.showerror("Ошибка запроса", f"Ошибка выполнения запроса:\n{str(e)}")
        except Exception as e:
            self.log_message(f"Непредвиденная ошибка: {str(e)}")
            messagebox.showerror("Ошибка", f"Непредвиденная ошибка:\n{str(e)}")
        finally:
            if self.connection:
                try:
                    self.connection.close()
                    self.log_message("Соединение с БД закрыто")
                except:
                    pass
            
            self.root.after(0, self.search_complete)
    
    def display_batch_results(self, batch):
        for row in batch:
            # Вставляем без серии и номера билета
            self.results_tree.insert('', tk.END, values=(
                row['Дата регистрации на терминале'],
                row['Дата списания'],
                row['Номер карты'],
                row['Стоимость проезда'],
                row['Стоимость билета'],
                row['Скидка'],
                row['С/Н терминала'],
                row['ГРЗ'],
                row['Маршрут'],
                row['Компания']
            ))
        
        self.root.update_idletasks()
    
    def display_results(self, results):
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
        
        for row in results:
            # Вставляем без серии и номера билета
            self.results_tree.insert('', tk.END, values=(
                row['Дата регистрации на терминале'],
                row['Дата списания'],
                row['Номер карты'],
                row['Стоимость проезда'],
                row['Стоимость билета'],
                row['Скидка'],
                row['С/Н терминала'],
                row['ГРЗ'],
                row['Маршрут'],
                row['Компания']
            ))
    
    def start_search(self):
        is_valid, error_message = self.validate_inputs()
        if not is_valid:
            messagebox.showwarning("Ошибка", error_message)
            return
        
        region_info = self.get_region_info()
        if not region_info:
            messagebox.showwarning("Ошибка", "Выберите регион")
            return
        
        self.clear_results()
        self.current_results = []
        self.row_count = 0
        
        self.stop_search_flag = False
        
        self.search_btn.config(state='disabled')
        self.stop_btn.config(state='normal')
        self.clear_params_btn.config(state='disabled')
        self.clear_results_btn.config(state='disabled')
        self.export_excel_btn.config(state='disabled')
        self.export_all_pdf_btn.config(state='disabled')
        
        self.progress_bar.start()
        self.log_message("Начало поиска...")
        
        self.search_thread = threading.Thread(target=self.execute_search, daemon=True)
        self.search_thread.start()
        
        self.log_message("Запуск поиска...")
    
    def stop_search(self):
        self.stop_search_flag = True
        self.stop_btn.config(state='disabled')
        self.log_message("Остановка поиска...")
    
    def search_complete(self):
        self.progress_bar.set_progress(100)
        
        self.search_btn.config(state='normal')
        self.stop_btn.config(state='disabled')
        self.clear_params_btn.config(state='normal')
        self.clear_results_btn.config(state='normal')
        
        if self.current_results:
            self.export_excel_btn.config(state='normal')
            self.export_all_pdf_btn.config(state='normal')
            self.log_message(f"Поиск завершен. Найдено {len(self.current_results)} записей")
        else:
            self.export_excel_btn.config(state='disabled')
            self.export_all_pdf_btn.config(state='disabled')
            self.log_message("Поиск завершен. Записей не найдено")
    
    def clear_search_params(self):
        self.region_var.set('')
        
        self.search_by_transaction.set(True)
        self.search_by_bankdate.set(False)
        self.toggle_date_frames()
        
        today = datetime.now()
        self.date_from_picker.set_date(today)
        self.date_to_picker.set_date(today)
        
        self.hour_from_combo.set("00")
        self.minute_from_combo.set("00")
        self.second_from_combo.set("00")
        
        self.hour_to_combo.set("23")
        self.minute_to_combo.set("59")
        self.second_to_combo.set("59")
        
        self.bankdate_from_picker.set_date(today)
        self.bankdate_to_picker.set_date(today)
        
        self.bank_hour_from_combo.set("00")
        self.bank_minute_from_combo.set("00")
        self.bank_second_from_combo.set("00")
        
        self.bank_hour_to_combo.set("23")
        self.bank_minute_to_combo.set("59")
        self.bank_second_to_combo.set("59")
        
        self.card_mask_var.set('')
        self.log_message("Параметры поиска очищены")
        self.progress_bar.set_progress(0)
    
    def clear_results(self):
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
        self.current_results = []
        self.export_excel_btn.config(state='disabled')
        self.export_all_pdf_btn.config(state='disabled')
        self.log_message("Результаты поиска очищены")
        self.progress_bar.set_progress(0)
    
    def export_to_excel(self):
        if not self.current_results:
            messagebox.showwarning("Нет данных", "Нет данных для экспорта")
            return
        
        try:
            date_from_str, date_to_str = self.get_datetime_from_ui() if self.search_by_transaction.get() else ("Не выбрано", "Не выбрано")
            bank_from_str, bank_to_str = self.get_bankdatetime_from_ui() if self.search_by_bankdate.get() else ("Не выбрано", "Не выбрано")
            
            region_text = self.region_var.get()
            region_name = region_text.split(' - ')[1] if ' - ' in region_text else "Регион"
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_region_name = "".join(c for c in region_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
            default_filename = f"Транзакции_{safe_region_name}_{timestamp}.xlsx"
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[
                    ("Excel файлы", "*.xlsx"),
                    ("Все файлы", "*.*")
                ],
                initialfile=default_filename
            )
            
            if not file_path:
                return
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Транзакции"
            
            headers = list(self.current_results[0].keys())
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for row_num, row_data in enumerate(self.current_results, 2):
                for col_num, header in enumerate(headers, 1):
                    value = row_data[header]
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            ws_meta = wb.create_sheet("Метаданные")
            
            meta_data = [
                ["Параметр", "Значение"],
                ["=" * 50, "=" * 50],
                ["Всего записей", len(self.current_results)],
                ["Дата экспорта", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ["Регион", region_name],
                ["Поиск по дате транзакции", "Да" if self.search_by_transaction.get() else "Нет"],
                ["Поиск по дате списания", "Да" if self.search_by_bankdate.get() else "Нет"],
            ]
            
            if self.search_by_transaction.get():
                meta_data.append(["Период регистрации", f"{date_from_str} - {date_to_str}"])
            
            if self.search_by_bankdate.get():
                meta_data.append(["Период списания", f"{bank_from_str} - {bank_to_str}"])
            
            meta_data.append(["Маска/номер карты", self.card_mask_var.get()])
            meta_data.append(["=" * 50, "=" * 50])
            
            for row_num, row_data in enumerate(meta_data, 1):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws_meta.cell(row=row_num, column=col_num, value=value)
                    if row_num == 1:
                        cell.font = Font(bold=True)
            
            for ws_sheet in [ws, ws_meta]:
                for col in ws_sheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_sheet.column_dimensions[column].width = adjusted_width
            
            wb.save(file_path)
            
            self.log_message(f"Результаты экспортированы в Excel файл: {file_path}")
            
            open_file = messagebox.askyesno(
                "Экспорт завершен", 
                f"Данные успешно экспортированы в файл:\n{file_path}\n\nОткрыть файл?"
            )
            
            if open_file:
                if os.name == 'nt':
                    os.startfile(file_path)
                elif os.name == 'posix':
                    import subprocess
                    subprocess.run(['xdg-open', file_path])
            
        except ImportError:
            error_msg = "Для экспорта в Excel требуется установить библиотеку openpyxl.\n"
            error_msg += "Установите командой: pip install openpyxl"
            self.log_message(f"Ошибка: {error_msg}")
            messagebox.showerror("Ошибка", error_msg)
        except PermissionError:
            error_msg = "Ошибка доступа к файлу. Возможно, файл открыт в другой программе."
            self.log_message(f"Ошибка экспорта: {error_msg}")
            messagebox.showerror("Ошибка экспорта", error_msg)
        except Exception as e:
            error_msg = f"Ошибка при экспорте в Excel: {str(e)}"
            self.log_message(f"Ошибка экспорта: {error_msg}")
            messagebox.showerror("Ошибка экспорта", error_msg)
    
    def export_all_to_pdf(self):
        """Экспорт всех найденных транзакций в один PDF файл с поддержкой переноса длинного текста"""
        if not self.current_results:
            messagebox.showwarning("Нет данных", "Нет данных для экспорта")
            return
        
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import os
            
            # Регистрируем шрифт с поддержкой кириллицы
            font_paths = [
                "C:/Windows/Fonts/arial.ttf",
                "C:/Windows/Fonts/times.ttf",
                "C:/Windows/Fonts/DejaVuSans.ttf",
            ]
            
            font_registered = False
            for font_path in font_paths:
                if os.path.exists(font_path):
                    try:
                        pdfmetrics.registerFont(TTFont('RussianFont', font_path))
                        font_registered = True
                        break
                    except:
                        continue
            
            if not font_registered:
                font_name = 'Helvetica'
            else:
                font_name = 'RussianFont'
            
            # Создаем имя файла
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            region_text = self.region_var.get()
            region_name = region_text.split(' - ')[1] if ' - ' in region_text else "Регион"
            safe_region_name = "".join(c for c in region_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
            default_filename = f"Поездки_{safe_region_name}_{timestamp}.pdf"
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[
                    ("PDF файлы", "*.pdf"),
                    ("Все файлы", "*.*")
                ],
                initialfile=default_filename
            )
            
            if not file_path:
                return
            
            # Создаем PDF документ
            doc = SimpleDocTemplate(file_path, pagesize=A4,
                                   rightMargin=72, leftMargin=72,
                                   topMargin=72, bottomMargin=72)
            story = []
            
            # Создаем стили
            styles = getSampleStyleSheet()
            
            # Стиль для ячеек с переносом текста
            cell_style = ParagraphStyle(
                'CellStyle',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=10,
                leading=14,
                alignment=0
            )
            
            # Общий заголовок
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Title'],
                fontName=font_name,
                fontSize=16,
                spaceAfter=30,
                textColor=colors.HexColor('#333333'),
                alignment=1
            )
            
            # Заголовок для каждой транзакции
            subtitle_style = ParagraphStyle(
                'CustomSubTitle',
                parent=styles['Heading2'],
                fontName=font_name,
                fontSize=14,
                spaceAfter=20,
                textColor=colors.HexColor('#0066cc'),
                alignment=0
            )
            
            # Форматируем стоимость
            def format_price(value):
                if value is None or value == '':
                    return "Не указано"
                try:
                    num_value = float(value)
                    return f"{num_value} ₽"
                except:
                    return str(value)
            
            # Форматируем скидку
            def format_discount(value):
                if value is None or value == '':
                    return "Не указано"
                try:
                    num_value = float(value)
                    if num_value == 0:
                        return "Не применимо"
                    return f"{num_value} ₽"
                except:
                    return str(value)
            
            # Маскируем номер карты
            def mask_card_number(card_number):
                if not card_number or card_number == 'Не указано' or card_number is None:
                    return 'Не указано'
                card_str = str(card_number).replace(' ', '')
                if len(card_str) >= 10:
                    first_six = card_str[:6]
                    last_four = card_str[-4:]
                    stars = '*' * (len(card_str) - 10)
                    return f"{first_six}{stars}{last_four}"
                return card_str
            
            # Функция для форматирования даты и времени
            def format_datetime(dt_str):
                if not dt_str or dt_str == 'Не указано' or dt_str is None:
                    return 'Не указано'
                try:
                    # Пробуем распарсить дату-время
                    str_dt = str(dt_str)
                    if len(str_dt) >= 19:
                        # Формат: YYYY-MM-DD HH:MM:SS
                        date_part = str_dt[:10]
                        time_part = str_dt[11:19]
                        # Преобразуем дату в формат ДД.ММ.ГГГГ
                        parts = date_part.split('-')
                        if len(parts) == 3:
                            formatted_date = f"{parts[2]}.{parts[1]}.{parts[0]}"
                            return f"{formatted_date} {time_part}"
                    elif len(str_dt) >= 10:
                        # Только дата
                        date_part = str_dt[:10]
                        parts = date_part.split('-')
                        if len(parts) == 3:
                            return f"{parts[2]}.{parts[1]}.{parts[0]}"
                    return str_dt
                except:
                    return str(dt_str)
            
            # Функция для автоматического переноса длинного текста
            def wrap_text(text, max_chars=35):
                """Разбивает длинный текст на строки с переносами"""
                text = str(text) if text else 'Не указано'
                if len(text) <= max_chars:
                    return text
                
                # Разбиваем по словам
                words = text.split()
                lines = []
                current_line = []
                current_length = 0
                
                for word in words:
                    word_len = len(word)
                    if current_length + word_len + (1 if current_line else 0) <= max_chars:
                        current_line.append(word)
                        current_length += word_len + (1 if current_line else 0)
                    else:
                        if current_line:
                            lines.append(' '.join(current_line))
                        current_line = [word]
                        current_length = word_len
                
                if current_line:
                    lines.append(' '.join(current_line))
                
                return '<br/>'.join(lines)
            
            # Получаем URL для проверки билета
            check_url = self.get_check_url()
            
            # Добавляем каждую транзакцию
            for i, transaction in enumerate(self.current_results):
                if i > 0:
                    story.append(PageBreak())
                
                # Получаем и форматируем дату поездки для заголовка
                trip_date = transaction.get('Дата регистрации на терминале', '')
                formatted_date = format_datetime(trip_date)
                
                # Заголовок с датой
                subtitle = Paragraph(f"Поездка от {formatted_date}", subtitle_style)
                story.append(subtitle)
                story.append(Spacer(1, 10))
                
                # Получаем название компании и применяем перенос
                company_name = transaction.get('Компания', 'Не указано')
                wrapped_company = wrap_text(company_name, 35)
                
                # Подготавливаем данные для таблицы - используем Paragraph для всех ячеек
                data = [
                    [Paragraph("<b>Параметр</b>", cell_style), 
                     Paragraph("<b>Значение</b>", cell_style)],
                    [Paragraph("Номер карты:", cell_style), 
                     Paragraph(mask_card_number(transaction.get('Номер карты', 'Не указано')), cell_style)],
                    [Paragraph("Дата и время списания:", cell_style), 
                     Paragraph(format_datetime(transaction.get('Дата списания', 'Не указано')), cell_style)],
                    [Paragraph("Дата и время поездки:", cell_style), 
                     Paragraph(format_datetime(transaction.get('Дата регистрации на терминале', 'Не указано')), cell_style)],
                    [Paragraph("Серия билета:", cell_style), 
                     Paragraph(str(transaction.get('Серия билета', 'Не указано')), cell_style)],
                    [Paragraph("Номер билета:", cell_style), 
                     Paragraph(str(transaction.get('Номер билета', 'Не указано')), cell_style)],
                    [Paragraph("Гос.номер:", cell_style), 
                     Paragraph(str(transaction.get('ГРЗ', 'Не указано')), cell_style)],
                    [Paragraph("Стоимость проезда:", cell_style), 
                     Paragraph(format_price(transaction.get('Стоимость проезда')), cell_style)],
                    [Paragraph("Скидка:", cell_style), 
                     Paragraph(format_discount(transaction.get('Скидка')), cell_style)],
                    [Paragraph("Компания-перевозчик:", cell_style), 
                     Paragraph(wrapped_company, cell_style)],
                    [Paragraph("<font color='blue'><u>Проверить билет:</u></font>", cell_style), 
                     Paragraph(f"<font color='blue'><u>{check_url}</u></font>", cell_style)],
                ]
                
                # Создаем таблицу
                table = Table(data, colWidths=[120*mm, 80*mm])
                
                # Стиль таблицы
                table_style = TableStyle([
                    ('BACKGROUND', (0, 0), (1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), font_name),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('WORDWRAP', (0, 1), (-1, -1), True),
                ])
                
                table.setStyle(table_style)
                story.append(table)
                story.append(Spacer(1, 20))
            
            # Добавляем дату генерации в конец
            date_style = ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontName=font_name,
                alignment=2,
                fontSize=8,
                textColor=colors.grey
            )
            gen_date = Paragraph(f"Отчет сформирован {datetime.now().strftime('%d.%m.%Y в %H:%M')}", date_style)
            story.append(gen_date)
            
            # Генерируем PDF
            doc.build(story)
            
            self.log_message(f"Экспорт завершен. Создан файл: {file_path}")
            
            # Предлагаем открыть файл
            open_file = messagebox.askyesno(
                "Экспорт завершен", 
                f"Экспортировано {len(self.current_results)} поездок в файл:\n{file_path}\n\nОткрыть файл?"
            )
            
            if open_file:
                if os.name == 'nt':
                    os.startfile(file_path)
                elif os.name == 'posix':
                    import subprocess
                    subprocess.run(['xdg-open', file_path])
                    
        except ImportError:
            messagebox.showerror("Ошибка", "Для экспорта в PDF требуется установить библиотеку reportlab.\nУстановите командой: pip install reportlab")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось создать PDF: {str(e)}")
    
    def on_closing(self):
        try:
            self.stop_search_flag = True
            self.root.destroy()
        except:
            try:
                self.root.quit()
            except:
                pass